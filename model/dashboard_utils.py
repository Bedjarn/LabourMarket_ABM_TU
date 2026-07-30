from __future__ import annotations

import hashlib
import json
import os
import shutil
import stat
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Callable, Dict, Iterable, Optional

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.lines import Line2D
import numpy as np
import pandas as pd

from base_parameters import BASE_PARAMS
from labour_market_model import LabourMarketModel, Producer


N_STEPS_DEFAULT = 1000

DEFAULT_DASHBOARD_MODES = ["npv_mean_field"]


def _preview_productivity_human(
    params: Dict[str, float],
    task_type: str,
    complexity_index: int,
    skill_level: str,
) -> float:
    """Evaluate the dashboard preview with Producer.productivity_human."""
    producer = SimpleNamespace(
        model=SimpleNamespace(
            a_prod=params["a_prod"],
            xi_prod=params["xi_prod"],
            phi_base=params["phi_base"],
            phi_decay=params["phi_decay"],
            nr_multiplier=params["nr_multiplier"],
        )
    )
    task = SimpleNamespace(
        task_type=task_type,
        complexity_index=complexity_index,
    )
    return Producer.productivity_human(producer, task, skill_level)


def _preview_productivity_ai(
    params: Dict[str, float],
    task_type: str,
    complexity_index: int,
) -> float:
    """Evaluate the dashboard preview with Producer.productivity_ai."""
    producer = SimpleNamespace(
        model=SimpleNamespace(
            phi_base=params["phi_base"],
            phi_decay=params["phi_decay"],
            nr_multiplier=params["nr_multiplier"],
        )
    )
    task = SimpleNamespace(
        task_type=task_type,
        complexity_index=complexity_index,
    )
    return Producer.productivity_ai(producer, task)

NUMERIC_PARAMS = [
    key for key, value in BASE_PARAMS.items()
    if isinstance(value, (int, float)) and not isinstance(value, bool)
]
BOOL_PARAMS = [key for key, value in BASE_PARAMS.items() if isinstance(value, bool)]
OFAT_PARAMS = [key for key in NUMERIC_PARAMS if key != "seed"] + BOOL_PARAMS

LEGACY_MODE_ALIASES = {"npv_rational": "npv_mean_field"}


def normalize_mode(mode: str) -> str:
    return LEGACY_MODE_ALIASES.get(mode, mode)


def normalize_modes(modes: Iterable[str]) -> list[str]:
    return [normalize_mode(mode) for mode in modes]


MODES = ["ulc", "npv_naive", "npv_adaptive", "npv_mean_field"]
MODE_LABELS = {
    "ulc": "ULC (baseline)",
    "npv_naive": "NPV - naive expectations",
    "npv_adaptive": "NPV - adaptive expectations",
    "npv_mean_field": "NPV - mean-field expectations",
    "npv_rational": "NPV - mean-field expectations",
}
MODE_COLORS = {
    "ulc": "#2d2d2d",
    "npv_naive": "#1f77b4",
    "npv_adaptive": "#ff7f0e",
    "npv_mean_field": "#2ca02c",
    "npv_rational": "#2ca02c",
}
MODE_LS = {
    "ulc": "-",
    "npv_naive": "--",
    "npv_adaptive": "-.",
    "npv_mean_field": ":",
    "npv_rational": ":",
}
MODE_MARKERS = {
    "ulc": "o",
    "npv_naive": "s",
    "npv_adaptive": "^",
    "npv_mean_field": "D",
    "npv_rational": "D",
}
NPV_MODES = ["npv_naive", "npv_adaptive", "npv_mean_field"]
CUMULATIVE_CHANNEL_METRICS = {
    "new_automations_this_step",
    "new_reactive_automations_this_step",
    "new_proactive_automations_this_step",
    "new_displaced_workers_this_step",
}
NPV_PLOT_METRICS = {
    "avg_npv_routine",
    "avg_npv_nonroutine",
    "share_adoptable_npv_routine",
    "share_adoptable_npv_nonroutine",
    "avg_realized_npv_this_step",
    "avg_realized_npv_cumulative",
}
NON_HISTOGRAM_FLOW_METRICS = {
    "reactive_automation_share_this_step",
    "proactive_automation_share_this_step",
}
OTHER_PLOT_GROUPS = {
    "Cost diagnostics": {
        "description": "These plots compare the running cost of AI and human inputs. Read them as input-side economics, not as final macro outcomes.",
        "metrics": [
            "input_cost_ai_routine",
            "input_cost_ai_nonroutine",
            "input_cost_human_high_routine",
            "input_cost_human_low_routine",
            "input_cost_human_high_nonroutine",
            "input_cost_human_low_nonroutine",
            "Avg labour cost per producer",
            "wage_bill",
        ],
    },
    "Flow diagnostics": {
        "description": "These are event-flow plots. They are naturally spiky because they count discrete automation and displacement events per step.",
        "metrics": [
            "new_automations_this_step",
            "new_reactive_automations_this_step",
            "new_proactive_automations_this_step",
        ],
    },
}
KEY_METRICS = [
    "ai_adoption_rate",
    "employment_rate_high",
    "employment_rate_low",
    "wage_high",
    "wage_low",
    "gini_income",
    "skill_wage_premium",
    "total_output",
    "price",
    "labour_share",
    "ULC",
    "wage_bill",
    "k_ai_current",
    "Avg profit per producer",
    "n_gate_blocks",
]
METRIC_LABELS = {
    "ai_adoption_rate": "Share of tasks automated",
    "new_automations_this_step": "New automation decisions",
    "new_reactive_automations_this_step": "Automations after worker exits",
    "new_proactive_automations_this_step": "Automations during expansion",
    "new_displaced_workers_this_step": "Workers displaced by automation",
    "employment_rate_high": "High-skill employment rate",
    "employment_rate_low": "Low-skill employment rate",
    "wage_high": "High-skill wage",
    "wage_low": "Low-skill wage",
    "gini_income": "Gini index of worker income",
    "skill_wage_premium": "Skill wage premium (w_h / w_l)",
    "total_output": "Total production",
    "price": "Goods market price",
    "labour_share": "Labour share of income",
    "ULC": "Unit labour cost",
    "wage_bill": "Total wage bill",
    "wage_bill_lagged": "Lagged wage bill",
    "profit_income_lagged": "Lagged profit income",
    "wage_demand_contribution": "Wage-led contribution to A",
    "profit_demand_contribution": "Profit-led contribution to A",
    "k_ai_current": "AI rental cost",
    "A_demand": "Demand level",
    "avg_npv_routine": "Average NPV of remaining routine tasks",
    "avg_npv_nonroutine": "Average NPV of remaining non-routine tasks",
    "share_adoptable_npv_routine": "Share of routine tasks still worth automating",
    "share_adoptable_npv_nonroutine": "Share of non-routine tasks still worth automating",
    "avg_realized_npv_this_step": "Average NPV of automations this step",
    "avg_realized_npv_cumulative": "Average NPV of all automations so far",
    "Avg profit per producer": "Average profit per producer",
    "Avg labour cost per producer": "Average labour cost per producer",
    "input_cost_ai_routine": "AI unit cost on routine tasks",
    "input_cost_ai_nonroutine": "AI unit cost on non-routine tasks",
    "input_cost_human_high_routine": "High-skill labour cost on routine tasks",
    "input_cost_human_low_routine": "Low-skill labour cost on routine tasks",
    "input_cost_human_high_nonroutine": "High-skill labour cost on non-routine tasks",
    "input_cost_human_low_nonroutine": "Low-skill labour cost on non-routine tasks",
    "reactive_automation_share_this_step": "Reactive automation share this step",
    "proactive_automation_share_this_step": "Proactive automation share this step",
    "w_min": "Minimum wage floor",
    "share_vast_high": "Share of high-skill workers on permanent contract",
    "share_vast_low": "Share of low-skill workers on permanent contract",
    "share_flex_high": "Share of high-skill workers on flex contract",
    "share_flex_low": "Share of low-skill workers on flex contract",
    "avg_tenure_years": "Average worker tenure (years)",
    "avg_tenure_vast_years": "Average tenure – permanent workers (years)",
    "avg_tenure_flex_years": "Average tenure – flex workers (years)",
    "avg_severance_per_task": "Average severance cost per task",
    "conversions_this_step": "Flex→vast conversions this step",
    "non_renewals_this_step": "Flex contract non-renewals this step",
    "n_gate_blocks": "Keynesian expansion gate blocks per step",
    "n_human_cheaper_tasks": "Tasks where cheapest human beats AI",
}
PLOT_METRICS = [
    metric for metric in METRIC_LABELS
    if metric not in NPV_PLOT_METRICS and metric not in NON_HISTOGRAM_FLOW_METRICS
]

# Single-metric panels: (metric, ylabel, title, ylim, pct)
# Dual-metric panels use a dict to signal combined plotting.
# Employment rates and wages are merged into one panel each so high and low
# can be compared directly.  Color = skill level; linestyle = adoption mode.
SKILL_COLORS = {"high": "#2563eb", "low": "#f97316"}  # blue=high, orange=low

LABOUR_MARKET_PANELS = [
    ("ai_adoption_rate", "Share of tasks", "Share of Tasks Automated", (0, 1), True),
    ("k_ai_current", "AI rental cost", "AI Rental Cost Over Time", None, False),
    # dual panels - dict signals combined treatment
    {"metrics": ["employment_rate_high", "employment_rate_low"],
     "skill_labels": {"employment_rate_high": "High-skill", "employment_rate_low": "Low-skill"},
     "ylabel": "Employment rate", "title": "Employment Rates by Skill Group", "ylim": (0, 1.05), "pct": True},
    {"metrics": ["wage_high", "wage_low"],
     "skill_labels": {"wage_high": "High-skill wage", "wage_low": "Low-skill wage"},
     "ylabel": "Wage", "title": "Wages by Skill Group", "ylim": None, "pct": False},
    ("gini_income", "Gini index", "Income Inequality Across Workers", (0, 1), False),
    ("skill_wage_premium", "w_h / w_l", "Skill Wage Premium", None, False),
    ("total_output", "Total output", "Total Production", None, False),
    ("price", "Goods market price", "Goods Market Price", None, False),
]

MACRO_PANELS = [
    ("labour_share", "Labour share", "Labour Share of Income", (0, 1), True, MODES),
    ("ULC", "Unit labour cost", "Unit Labour Cost", None, False, MODES),
    ("wage_bill", "Total wage bill", "Total Wage Bill", None, False, MODES),
    ("A_demand", "Demand level", "Demand Level", None, False, MODES),
    # Kaleckian factor-income decomposition: shows the labour vs. capital income
    # streams that BOTH feed into demand A. With AI automation, wage income falls
    # while profit income rises — the two series should display the classic
    # functional-income redistribution. (Per supervisor feedback May 2026.)
    {"metrics": ["wage_bill_lagged", "profit_income_lagged"],
     "series_labels": {
         "wage_bill_lagged":     "Wage income (W)",
         "profit_income_lagged": "Profit income (Pi)",
     },
     "series_colors": {
         "wage_bill_lagged":     "#2563eb",  # blue  = labour
         "profit_income_lagged": "#16a34a",  # green = capital
     },
     "ylabel": "Aggregate factor income", "title": "Functional Income Distribution (Wage vs Profit)",
     "ylim": None, "pct": False, "modes": MODES},
    # Decomposition of the demand shifter A into its wage-led and profit-led
    # parts: A - A_base = gamma * W + gamma_pi * Pi. Lets the user inspect
    # whether demand resilience comes from the labour or the capital channel.
    {"metrics": ["wage_demand_contribution", "profit_demand_contribution"],
     "series_labels": {
         "wage_demand_contribution":   "Wage channel (gamma * W)",
         "profit_demand_contribution": "Profit channel (gamma_pi * Pi)",
     },
     "series_colors": {
         "wage_demand_contribution":   "#2563eb",
         "profit_demand_contribution": "#16a34a",
     },
     "ylabel": "Contribution to A", "title": "Demand Decomposition: Wage-led vs Profit-led Channel",
     "ylim": None, "pct": False, "modes": MODES},
    ("Avg profit per producer", "Profit", "Average Profit per Producer", None, False, MODES),
    ("n_gate_blocks", "Gate blocks per step", "Keynesian Expansion Gate: Blocks per Step", None, False, MODES),
    {"metrics": ["new_reactive_automations_this_step", "new_proactive_automations_this_step"],
     "series_labels": {
         "new_reactive_automations_this_step": "Reactive channel",
         "new_proactive_automations_this_step": "Proactive channel",
     },
     "series_colors": {
         "new_reactive_automations_this_step": "#2563eb",
         "new_proactive_automations_this_step": "#16a34a",
     },
     "ylabel": "Automations per step", "title": "How AI Adoption Happens: Replacement vs Expansion", "ylim": None, "pct": False, "modes": MODES, "style": "bar"},
]

EMPLOYMENT_PROTECTION_PANELS = [
    # Vast/flex share panels by skill group
    {"metrics": ["share_vast_high", "share_vast_low"],
     "skill_labels": {"share_vast_high": "High-skill (vast)", "share_vast_low": "Low-skill (vast)"},
     "ylabel": "Share on permanent contract", "title": "Permanent Contract Share by Skill Group",
     "ylim": (0, 1.05), "pct": True},
    {"metrics": ["share_flex_high", "share_flex_low"],
     "skill_labels": {"share_flex_high": "High-skill (flex)", "share_flex_low": "Low-skill (flex)"},
     "ylabel": "Share on flex contract", "title": "Flex Contract Share by Skill Group",
     "ylim": (0, 1.05), "pct": True},
    # Tenure panels
    {"metrics": ["avg_tenure_vast_years", "avg_tenure_flex_years"],
     "skill_labels": {"avg_tenure_vast_years": "Permanent workers", "avg_tenure_flex_years": "Flex workers"},
     "ylabel": "Average tenure (years)", "title": "Tenure by Contract Type",
     "ylim": None, "pct": False},
    ("avg_tenure_years", "Average tenure (years)", "Overall Average Tenure", None, False),
    # Flow metrics: conversions and non-renewals
    {"metrics": ["conversions_this_step", "non_renewals_this_step"],
     "skill_labels": {"conversions_this_step": "Flex→vast conversions", "non_renewals_this_step": "Non-renewals"},
     "ylabel": "Workers per step", "title": "Chain-Clause Events per Step",
     "ylim": None, "pct": False},
    # Severance cost
    ("avg_severance_per_task", "Severance cost", "Average Severance Cost per Task", None, False),
]


@dataclass
class RunBundle:
    results: Dict[str, pd.DataFrame]
    models: Dict[str, LabourMarketModel]
    params: Dict[str, float]
    n_steps: int
    modes: list[str]
    run_label: str
    run_id: Optional[str] = None


def ensure_derived_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Add dashboard metrics that can be recovered from existing time-series."""
    if "skill_wage_premium" not in df.columns and {"wage_high", "wage_low"}.issubset(df.columns):
        wage_high = pd.to_numeric(df["wage_high"], errors="coerce")
        wage_low = pd.to_numeric(df["wage_low"], errors="coerce")
        df["skill_wage_premium"] = wage_high.div(wage_low.where(wage_low > 0, np.nan))
    return df


def run_simulation(
    params: Optional[Dict[str, float]] = None,
    n_steps: int = N_STEPS_DEFAULT,
    modes: Optional[Iterable[str]] = None,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
    run_label: str = "",
    log_decisions: bool = False,
) -> RunBundle:
    resolved_params = {**BASE_PARAMS, **(params or {})}
    if "nr_penalty" in resolved_params and "nr_multiplier" not in resolved_params:
        resolved_params["nr_multiplier"] = resolved_params.pop("nr_penalty")
    selected_modes = normalize_modes(list(modes or MODES))
    results: Dict[str, pd.DataFrame] = {}
    models: Dict[str, LabourMarketModel] = {}

    total_ticks = len(selected_modes) * max(1, n_steps)
    completed_ticks = 0
    for mode in selected_modes:
        model_params = {**resolved_params, "adoption_mode": mode, "log_decisions": log_decisions}
        model = LabourMarketModel(**model_params)
        for step in range(n_steps):
            model.step()
            completed_ticks += 1
            if progress_callback:
                progress_callback(mode, completed_ticks, total_ticks)
        results[mode] = ensure_derived_metrics(model.datacollector.get_model_vars_dataframe())
        models[mode] = model

    return RunBundle(
        results=results,
        models=models,
        params=resolved_params,
        n_steps=n_steps,
        modes=selected_modes,
        run_label=run_label.strip(),
    )


def _apply_rolling(series: "pd.Series", window: int) -> "pd.Series":
    """Apply rolling average if window > 1, otherwise return unchanged."""
    if window > 1:
        return series.rolling(window, min_periods=1).mean()
    return series


def _plot_panel(ax, results: Dict[str, pd.DataFrame], metric: str, ylabel: str, title: str,
                ylim=None, modes=None, pct=False, rolling: int = 1):
    selected_modes = list(modes or results.keys())
    for mode in selected_modes:
        df = results.get(mode)
        if df is not None:
            df = ensure_derived_metrics(df)
        if df is None or metric not in df.columns:
            continue
        values = _apply_rolling(df[metric], rolling)
        ax.plot(
            df.index,
            values,
            label=MODE_LABELS.get(mode, mode),
            color=MODE_COLORS.get(mode),
            linestyle=MODE_LS.get(mode, "-"),
            linewidth=1.8,
        )
    _title = f"{title}  ({rolling}-step avg)" if rolling > 1 else title
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(_title, fontsize=10, fontweight="bold", pad=10)
    if ylim:
        ax.set_ylim(*ylim)
    if pct:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    ax.tick_params(axis="both", labelsize=8)
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=7.5, loc="best")


def _plot_combined_panel(ax, results: Dict[str, pd.DataFrame], panel: dict, rolling: int = 1):
    metrics = panel["metrics"]
    skill_labels = panel.get("skill_labels", {})
    series_labels = panel.get("series_labels", {})
    series_colors = panel.get("series_colors", {})
    style = panel.get("style", "line")
    color_by_mode = panel.get("color_by_mode", False)
    ylim = panel["ylim"]
    pct = panel["pct"]
    selected_modes = list(panel.get("modes", results.keys()))
    active_series: list[tuple[str, object, str, str]] = []
    for mode in selected_modes:
        df = results.get(mode)
        if df is not None:
            df = ensure_derived_metrics(df)
        if df is None:
            continue
        for metric in metrics:
            if metric not in df.columns:
                continue
            if color_by_mode:
                color = MODE_COLORS.get(mode, "#555")
                metric_label = series_labels.get(metric, skill_labels.get(metric, metric))
            elif series_colors:
                color = series_colors.get(metric, "#555")
                metric_label = series_labels.get(metric, metric)
            else:
                color = SKILL_COLORS.get(
                    "high" if "high" in metric else "low", "#555"
                )
                metric_label = skill_labels.get(metric, metric)
            label = f"{MODE_LABELS.get(mode, mode)} - {metric_label}"
            active_series.append((mode, df, metric, color))
            if style != "bar":
                values = _apply_rolling(df[metric], rolling)
                ax.plot(
                    df.index,
                    values,
                    label=label,
                    color=color,
                    linestyle=MODE_LS.get(mode, "-"),
                    marker=MODE_MARKERS.get(mode, None),
                    markersize=3.2,
                    markevery=max(1, len(df.index) // 14),
                    linewidth=1.8,
                )
    if style == "bar" and active_series:
        n_series = len(active_series)
        width = 0.8 / max(1, n_series)
        for idx, (mode, df, metric, color) in enumerate(active_series):
            metric_label = series_labels.get(metric, skill_labels.get(metric, metric))
            label = f"{MODE_LABELS.get(mode, mode)} - {metric_label}"
            offset = (idx - (n_series - 1) / 2) * width
            ax.bar(
                df.index + offset,
                df[metric],
                width=width,
                label=label,
                color=color,
                alpha=0.72,
                edgecolor="white",
                linewidth=0.2,
                align="center",
            )
    _ptitle = f"{panel['title']}  ({rolling}-step avg)" if rolling > 1 else panel["title"]
    ax.set_ylabel(panel["ylabel"], fontsize=9)
    ax.set_title(_ptitle, fontsize=10, fontweight="bold", pad=10)
    if ylim:
        ax.set_ylim(*ylim)
    if pct:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    ax.tick_params(axis="both", labelsize=8)
    ax.grid(True, linestyle="--", alpha=0.5)
    if style == "bar" or series_colors:
        ax.legend(fontsize=6.5, loc="best")
    else:
        skill_handles = [
            Line2D([0], [0], color=SKILL_COLORS["high"], lw=2.2, label="High-skill"),
            Line2D([0], [0], color=SKILL_COLORS["low"], lw=2.2, label="Low-skill"),
        ]
        mode_handles = [
            Line2D(
                [0], [0],
                color="#475569",
                lw=2.0,
                linestyle=MODE_LS.get(mode, "-"),
                marker=MODE_MARKERS.get(mode, None),
                markersize=4,
                label=MODE_LABELS.get(mode, mode),
            )
            for mode in selected_modes
        ]
        legend_skill = ax.legend(
            handles=skill_handles,
            title="Skill",
            fontsize=6.5,
            title_fontsize=7,
            loc="upper right",
            frameon=True,
        )
        ax.add_artist(legend_skill)
        ax.legend(
            handles=mode_handles,
            title="Mode",
            fontsize=6.3,
            title_fontsize=7,
            loc="lower left",
            frameon=True,
        )


def build_labour_market_figure(results: Dict[str, pd.DataFrame], rolling: int = 1):
    n_panels = len(LABOUR_MARKET_PANELS)
    ncols = 2
    nrows = (n_panels + 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(14, 4.5 * nrows))
    fig.suptitle("Labour Market ABM - Labour market and AI adoption", fontsize=13, y=0.995)

    for axis, panel in zip(axes.flatten(), LABOUR_MARKET_PANELS):
        if isinstance(panel, dict):
            _plot_combined_panel(axis, results, panel, rolling=rolling)
        else:
            metric, ylabel, title, ylim, pct = panel
            _plot_panel(axis, results, metric, ylabel, title, ylim=ylim, pct=pct, rolling=rolling)

    for ax in axes[-1, :]:
        ax.set_xlabel("Model step", fontsize=8)
    for ax in axes.flatten()[n_panels:]:
        ax.set_visible(False)
    fig.align_ylabels()
    fig.tight_layout(rect=(0, 0, 1, 0.965), h_pad=2.0, w_pad=1.2)
    return fig


def build_macro_figure(results: Dict[str, pd.DataFrame], rolling: int = 1):
    n_panels = len(MACRO_PANELS)
    ncols = 2
    nrows = (n_panels + 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(14, 4.5 * nrows))
    fig.suptitle("Labour Market ABM - Macro and financial indicators", fontsize=13, y=0.995)
    for axis, panel in zip(axes.flatten(), MACRO_PANELS):
        if isinstance(panel, dict):
            available_modes = [mode for mode in panel.get("modes", results.keys()) if mode in results]
            _plot_combined_panel(axis, results, {**panel, "modes": available_modes}, rolling=rolling)
        else:
            metric, ylabel, title, ylim, pct, modes = panel
            available_modes = [mode for mode in modes if mode in results]
            _plot_panel(axis, results, metric, ylabel, title, ylim=ylim, modes=available_modes, pct=pct, rolling=rolling)
    for ax in axes[-1, :]:
        ax.set_xlabel("Model step", fontsize=8)
    for ax in axes.flatten()[n_panels:]:
        ax.set_visible(False)
    fig.align_ylabels()
    fig.tight_layout(rect=(0, 0, 1, 0.965), h_pad=2.0, w_pad=1.2)
    return fig


def build_other_figure(results: Dict[str, pd.DataFrame]):
    return build_other_figures(results).get("All other indicators")


def _resolve_other_metrics(results: Dict[str, pd.DataFrame]):
    displayed_metrics = set()
    for panel in LABOUR_MARKET_PANELS:
        if isinstance(panel, dict):
            displayed_metrics.update(panel["metrics"])
        else:
            displayed_metrics.add(panel[0])
    for panel in MACRO_PANELS:
        if isinstance(panel, dict):
            displayed_metrics.update(panel["metrics"])
        else:
            displayed_metrics.add(panel[0])

    excluded_metrics = (displayed_metrics - CUMULATIVE_CHANNEL_METRICS) | {
        "adoption_mode",
        "ai_irreversible",
        "w_min",
        *NPV_PLOT_METRICS,
        *NON_HISTOGRAM_FLOW_METRICS,
    }
    all_metrics = []
    for df in results.values():
        for column in df.columns:
            if column in excluded_metrics:
                continue
            if column not in all_metrics:
                all_metrics.append(column)
    return all_metrics


def _metric_has_signal(results: Dict[str, pd.DataFrame], metric: str, tolerance: float = 1e-12) -> bool:
    saw_any = False
    for df in results.values():
        if metric not in df.columns:
            continue
        series = pd.to_numeric(df[metric], errors="coerce").dropna()
        if series.empty:
            continue
        saw_any = True
        if float(series.max() - series.min()) > tolerance:
            return True
        if float(series.abs().max()) > tolerance:
            return True
    return saw_any


def _build_metric_panels_figure(
    results: Dict[str, pd.DataFrame],
    metrics: list[str],
    title: str,
):
    if not metrics:
        return None

    n_panels = len(metrics)
    ncols = 2
    nrows = (n_panels + 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(14, 4.5 * nrows))
    fig.suptitle(title, fontsize=13, y=0.995)
    axes_flat = np.atleast_1d(axes).flatten()

    for axis, metric in zip(axes_flat, metrics):
        pct = metric in {"share_adoptable_npv_routine", "share_adoptable_npv_nonroutine"}
        if metric in CUMULATIVE_CHANNEL_METRICS:
            cumulative_results = {
                mode: df.assign(**{metric: pd.to_numeric(df[metric], errors="coerce").fillna(0).cumsum()})
                if metric in df.columns else df
                for mode, df in results.items()
            }
            _plot_combined_panel(
                axis,
                cumulative_results,
                {
                    "metrics": [metric],
                    "series_labels": {metric: METRIC_LABELS.get(metric, metric)},
                    "ylabel": f"Cumulative {METRIC_LABELS.get(metric, metric).lower()}",
                    "title": f"Cumulative {METRIC_LABELS.get(metric, metric)}",
                    "ylim": None,
                    "pct": False,
                    "modes": list(cumulative_results.keys()),
                    "style": "line",
                    "color_by_mode": True,
                },
            )
        else:
            _plot_panel(
                axis,
                results,
                metric,
                METRIC_LABELS.get(metric, metric),
                METRIC_LABELS.get(metric, metric),
                pct=pct,
            )

    for ax in axes_flat[-ncols:]:
        ax.set_xlabel("Model step", fontsize=8)
    for ax in axes_flat[n_panels:]:
        ax.set_visible(False)
    fig.align_ylabels()
    fig.tight_layout(rect=(0, 0, 1, 0.965), h_pad=2.0, w_pad=1.2)
    return fig


def build_other_figures(results: Dict[str, pd.DataFrame]):
    all_metrics = _resolve_other_metrics(results)
    if not all_metrics:
        return {}

    figures: dict[str, plt.Figure] = {}
    used_metrics: set[str] = set()
    for group_name, group in OTHER_PLOT_GROUPS.items():
        group_metrics = [
            metric
            for metric in group["metrics"]
            if metric in all_metrics and _metric_has_signal(results, metric)
        ]
        if not group_metrics:
            continue
        used_metrics.update(group_metrics)
        figure = _build_metric_panels_figure(
            results,
            group_metrics,
            f"Labour Market ABM - {group_name}",
        )
        if figure is not None:
            figures[group_name] = figure

    remaining_metrics = [
        metric for metric in all_metrics
        if metric not in used_metrics and _metric_has_signal(results, metric)
    ]
    repeated_flow_metrics = [
        metric for metric in (
            "new_reactive_automations_this_step",
            "new_proactive_automations_this_step",
        )
        if metric in all_metrics and _metric_has_signal(results, metric)
    ]
    remaining_metrics = repeated_flow_metrics + [
        metric for metric in remaining_metrics if metric not in repeated_flow_metrics
    ]
    if remaining_metrics:
        figure = _build_metric_panels_figure(
            results,
            remaining_metrics,
            "Labour Market ABM - Additional diagnostics",
        )
        if figure is not None:
            figures["Additional diagnostics"] = figure
    return figures


def build_employment_protection_figure(results: Dict[str, pd.DataFrame]) -> plt.Figure:
    """
    Employment-protection diagnostic figure.

    Layout (4 rows × 2 cols):
      Row 1: Permanent contract share – high-skill | Permanent contract share – low-skill
      Row 2: Average tenure (years)               | Average severance cost per task
      Row 3: Flex→Permanent conversions (full panel, wide)
      Row 4: Flex contract non-renewals (full panel, wide)

    The two flow panels (rows 3-4) use a two-layer rendering:
      - Light translucent raw series (step plot)
      - Bold rolling-average line on top for trend readability
    """
    # Adaptive smoothing: use 5% of total run length, clamped to [15, 60]
    _n_steps = max((len(df) for df in results.values() if len(df) > 0), default=1000)
    _SMOOTH = int(max(15, min(60, _n_steps * 0.05)))

    # Stock panels (rows 1-2, 2-col layout)
    _STOCK_PANELS = [
        ("share_vast_high",       "Share on permanent contract", "Permanent Contract Share — High-skill",  (0, 1.05), True),
        ("share_vast_low",        "Share on permanent contract", "Permanent Contract Share — Low-skill",   (0, 1.05), True),
        ("avg_tenure_years",      "Years",                       "Average Tenure (all workers)",           None,      False),
        ("avg_severance_per_task","€ (severance cost)",          "Average Severance Cost per Task",        None,      False),
    ]

    # Flow panels get their own full-width rows
    _FLOW_PANELS = [
        ("conversions_this_step",  "Flex → Permanent conversions per step",
         f"Flex → Permanent Contract Conversions  ({_SMOOTH}-step rolling avg)"),
        ("non_renewals_this_step", "Flex contract non-renewals per step",
         f"Flex Contract Non-renewals  ({_SMOOTH}-step rolling avg)"),
    ]

    fig = plt.figure(figsize=(15, 14))
    # Grid: 4 rows, 2 cols — first 2 rows use both cols, last 2 rows span full width
    gs = fig.add_gridspec(4, 2, hspace=0.42, wspace=0.30)

    stock_axes = [fig.add_subplot(gs[r, c]) for r in range(2) for c in range(2)]
    flow_axes  = [fig.add_subplot(gs[2, :]), fig.add_subplot(gs[3, :])]

    fig.suptitle("Employment Protection Diagnostics", fontsize=14, fontweight="bold", y=1.01)

    # ── Stock panels (shares, tenure, severance) ───────────────────────────────
    for ax, (metric, ylabel, title, ylim, pct) in zip(stock_axes, _STOCK_PANELS):
        plotted = False
        for mode in results:
            df = results[mode]
            if metric not in df.columns:
                continue
            series = df[metric].dropna()
            if series.empty:
                continue
            ax.plot(
                series.index, series,
                label=MODE_LABELS.get(mode, mode),
                color=MODE_COLORS.get(mode),
                linestyle=MODE_LS.get(mode, "-"),
                linewidth=1.8,
            )
            plotted = True
        ax.set_title(title, fontsize=10)
        ax.set_xlabel("Model step")
        ax.set_ylabel(ylabel)
        if ylim:
            ax.set_ylim(*ylim)
        if pct:
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0%}"))
        ax.grid(alpha=0.22)
        if plotted:
            ax.legend(fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data\n(employment_protection=False?)",
                    ha="center", va="center", transform=ax.transAxes, color="gray", fontsize=9)

    # ── Flow panels (conversions / non-renewals) — dual-layer rendering ────────
    for ax, (metric, ylabel, title) in zip(flow_axes, _FLOW_PANELS):
        plotted = False
        for mode in results:
            df = results[mode]
            if metric not in df.columns:
                continue
            raw = df[metric].dropna()
            if raw.empty:
                continue
            color = MODE_COLORS.get(mode, "#475569")
            # Layer 1: raw data as very light step plot
            ax.step(raw.index, raw, where="mid",
                    color=color, alpha=0.18, linewidth=0.7)
            # Layer 2: bold rolling average
            smooth = raw.rolling(_SMOOTH, min_periods=1).mean()
            ax.plot(smooth.index, smooth,
                    label=MODE_LABELS.get(mode, mode),
                    color=color,
                    linestyle=MODE_LS.get(mode, "-"),
                    linewidth=2.2)
            plotted = True

        ax.set_title(title, fontsize=11, fontweight="bold")
        ax.set_xlabel("Model step")
        ax.set_ylabel(ylabel)
        ax.set_ylim(bottom=0)
        ax.grid(alpha=0.22)
        ax.legend(fontsize=9, title=f"Bold line = {_SMOOTH}-step rolling avg  |  faint = raw")
        if not plotted:
            ax.text(0.5, 0.5, "No data\n(employment_protection=False?)",
                    ha="center", va="center", transform=ax.transAxes, color="gray", fontsize=9)

    fig.tight_layout()
    return fig


# ── Investment decision log ────────────────────────────────────────────────────

# Preferred column order for the Excel export (columns not in this list are
# appended at the end in the order they were collected).
_LOG_COLUMN_ORDER = [
    # Identity
    "step", "mode", "channel", "evaluated", "decision", "decision_reason",
    # Task
    "producer_id", "alpha", "task_id", "task_type", "complexity",
    # Market conditions
    "wage_h_now", "wage_l_now", "k_ai_now",
    # Productivities
    "prod_h_high", "prod_h_low", "prod_ai",
    # Unit costs now
    "uc_human_now", "uc_ai_now",
    # NPV breakdown
    "npv", "investment_cost", "severance_cost", "total_upfront",
    "hurdle_threshold", "npv_minus_hurdle", "discounted_savings",
    "uc_ai_at_horizon",
    # NPV settings
    "T_horizon", "r_discount", "eta_hurdle",
    # Expected wages
    "exp_wage_h_mean", "exp_wage_l_mean", "expectation_detail",
    # Employment protection
    "n_vast_on_task", "n_flex_on_task",
    "departed_contract", "departed_tenure_y",
    "severance_detail",
]

_LOG_COLUMN_LABELS = {
    "step":               "Step",
    "mode":               "Adoption mode",
    "channel":            "Channel (reactive/proactive)",
    "evaluated":          "p_evaluate fired?",
    "decision":           "Decision: automate?",
    "decision_reason":    "Reason",
    "producer_id":        "Producer ID",
    "alpha":              "α (routine share)",
    "task_id":            "Task ID",
    "task_type":          "Task type",
    "complexity":         "Complexity index",
    "wage_h_now":         "Wage high-skill (now)",
    "wage_l_now":         "Wage low-skill (now)",
    "k_ai_now":           "AI rental cost (now)",
    "prod_h_high":        "Productivity: human high-skill",
    "prod_h_low":         "Productivity: human low-skill",
    "prod_ai":            "Productivity: AI",
    "uc_human_now":       "Unit cost: cheapest human (now)",
    "uc_ai_now":          "Unit cost: AI (now)",
    "npv":                "NPV",
    "investment_cost":    "Investment cost I(x)",
    "severance_cost":     "Severance cost S(x)",
    "total_upfront":      "Total upfront cost I(x)+S(x)",
    "hurdle_threshold":   "Hurdle threshold η·I(x)",
    "npv_minus_hurdle":   "NPV − hurdle (>0 → automate)",
    "discounted_savings": "Discounted savings Σ ΔC/(1+r)^t",
    "uc_ai_at_horizon":   "Unit cost AI at end of horizon",
    "T_horizon":          "Planning horizon T",
    "r_discount":         "Discount rate r",
    "eta_hurdle":         "Hurdle rate η",
    "exp_wage_h_mean":    "Expected wage high-skill (mean over horizon)",
    "exp_wage_l_mean":    "Expected wage low-skill (mean over horizon)",
    "expectation_detail": "Expectation detail",
    "n_vast_on_task":     "Vast workers on task",
    "n_flex_on_task":     "Flex workers on task",
    "departed_contract":  "Departed worker contract",
    "departed_tenure_y":  "Departed worker tenure (years)",
    "severance_detail":   "Severance per vast worker",
}


def build_investment_log_dataframe(
    bundle: "RunBundle",
    experiment_name: str = "",
) -> pd.DataFrame:
    """
    Combine investment_log from all modes in a RunBundle into one DataFrame.
    Returns an empty DataFrame if logging was not enabled (no records).
    Optionally stamps an experiment_name column.
    """
    frames = []
    for mode in bundle.modes:
        model = bundle.models.get(mode)
        if model is None or not model.investment_log:
            continue
        df = pd.DataFrame(model.investment_log)
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    if experiment_name:
        combined.insert(0, "experiment", experiment_name)
    # Reorder columns: known columns first, then any extras
    present_known   = [c for c in _LOG_COLUMN_ORDER if c in combined.columns]
    extra_cols      = [c for c in combined.columns if c not in _LOG_COLUMN_ORDER
                       and c != "experiment"]
    ordered         = (["experiment"] if experiment_name else []) + present_known + extra_cols
    return combined[ordered]


def build_investment_log_excel_bytes(
    log_df: pd.DataFrame,
    experiment_logs: Optional[list[dict]] = None,
) -> bytes:
    """
    Build an Excel workbook from investment log data and return it as bytes
    suitable for st.download_button.

    log_df          : DataFrame from build_investment_log_dataframe() for a single bundle.
    experiment_logs : list of {"name": str, "df": pd.DataFrame} for multi-experiment export.

    Workbook layout
    ───────────────
    Sheet "All decisions"            – every row
    Sheet "Automated"                – decision == True
    Sheet "Evaluated – not automated"– evaluated == True AND decision == False
    Sheet "Not evaluated"            – evaluated == False
    Sheet "Per-step summary"         – counts grouped by step + mode
    [Optional] one sheet per experiment when experiment_logs is provided
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ── helper: rename columns to human-friendly labels ───────────────────────
    def _renamed(df: pd.DataFrame) -> pd.DataFrame:
        return df.rename(columns=_LOG_COLUMN_LABELS)

    # ── helper: write one DataFrame to a sheet with header formatting ─────────
    def _write_sheet(ws, df: pd.DataFrame, freeze: bool = True) -> None:
        if df.empty:
            ws.append(["(no data)"])
            return
        df_r = _renamed(df)
        for r_idx, row in enumerate(dataframe_to_rows(df_r, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:  # header row
                for cell in ws[r_idx]:
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor="1F4E79")
                    cell.alignment = Alignment(wrap_text=True)
        # Auto-width (capped at 50)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        if freeze:
            ws.freeze_panes = ws["A2"]

    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)   # delete default empty sheet

    # ── Sheet 1: All decisions ────────────────────────────────────────────────
    _write_sheet(wb.create_sheet("All decisions"), log_df)

    # ── Sheet 2: Automated ───────────────────────────────────────────────────
    auto_df = log_df[log_df.get("decision", pd.Series(dtype=bool)) == True]  # noqa: E712
    _write_sheet(wb.create_sheet("Automated"), auto_df)

    # ── Sheet 3: Evaluated – not automated ───────────────────────────────────
    rej_mask = (log_df.get("evaluated", pd.Series(dtype=bool)) == True) & \
               (log_df.get("decision",  pd.Series(dtype=bool)) == False)   # noqa: E712
    _write_sheet(wb.create_sheet("Evaluated – not automated"), log_df[rej_mask])

    # ── Sheet 4: Not evaluated ────────────────────────────────────────────────
    not_eval = log_df[log_df.get("evaluated", pd.Series(dtype=bool)) == False]  # noqa: E712
    _write_sheet(wb.create_sheet("Not evaluated"), not_eval)

    # ── Sheet 5: Per-step summary ─────────────────────────────────────────────
    if not log_df.empty and "step" in log_df.columns and "mode" in log_df.columns:
        grp_cols = ["step", "mode"]
        if "experiment" in log_df.columns:
            grp_cols = ["experiment"] + grp_cols
        summary = (
            log_df.groupby(grp_cols, dropna=False)
            .agg(
                total_evaluations=("decision", "count"),
                automated=("decision",  lambda s: (s == True).sum()),   # noqa: E712
                not_automated=("decision", lambda s: (s == False).sum()),  # noqa: E712
                npv_mean=("npv", "mean"),
                npv_min=("npv",  "min"),
                npv_max=("npv",  "max"),
            )
            .reset_index()
        )
        _write_sheet(wb.create_sheet("Per-step summary"), summary)

    # ── Optional: one sheet per experiment ────────────────────────────────────
    if experiment_logs:
        for item in experiment_logs:
            name = item["name"][:28]  # Excel sheet names max 31 chars
            _write_sheet(wb.create_sheet(name), item["df"])

    wb.save(buf)
    return buf.getvalue()


def build_summary_dataframe(bundle: RunBundle) -> pd.DataFrame:
    summary_rows = []
    for mode in bundle.modes:
        df = ensure_derived_metrics(bundle.results[mode])
        row = {"mode": MODE_LABELS.get(mode, mode)}
        for metric in KEY_METRICS:
            if metric not in df.columns:
                continue
            row[f"{metric}_mean"] = df[metric].mean()
            row[f"{metric}_final"] = df[metric].iloc[-1]
            row[f"{metric}_min"] = df[metric].min()
            row[f"{metric}_max"] = df[metric].max()
        summary_rows.append(row)
    return pd.DataFrame(summary_rows).set_index("mode")


def build_combined_dataframe(bundle: RunBundle) -> pd.DataFrame:
    frames = []
    for mode in bundle.modes:
        df = ensure_derived_metrics(bundle.results[mode]).copy()
        df.columns = [f"{mode}__{column}" for column in df.columns]
        frames.append(df)
    combined = pd.concat(frames, axis=1)
    combined.index.name = "step"
    return combined


def build_producer_dataframe(bundle: RunBundle) -> pd.DataFrame:
    if not bundle.models:
        return pd.DataFrame()
    if "ulc" in bundle.models:
        model = bundle.models["ulc"]
    else:
        model = bundle.models[bundle.modes[0]]
    rows = []
    for producer in model.producers:
        for task in producer.tasks:
            rows.append(
                {
                    "producer_id": producer.unique_id,
                    "alpha": round(producer.alpha, 4),
                    "task_id": task.task_id,
                    "task_type": task.task_type,
                    "complexity_index": task.complexity_index,
                    "automated": task.automated,
                    "n_ai": task.n_ai,
                    "n_workers": len(task.employees),
                    "worker_skills": ",".join(worker.skill_level for worker in task.employees),
                }
            )
    return pd.DataFrame(rows)


def build_timing_dataframe(bundle: RunBundle) -> pd.DataFrame:
    if not bundle.models:
        return pd.DataFrame(columns=["mode", "producer_id", "alpha", "task_type", "complexity_index", "investment_step"])
    timing_rows = []
    for mode in [mode for mode in bundle.modes if mode in NPV_MODES and mode in bundle.models]:
        model = bundle.models[mode]
        for producer in model.producers:
            for task in producer.tasks:
                original = task.ai_original_step if task.ai_original_step is not None else task.ai_investment_step
                if task.automated and original is not None:
                    timing_rows.append(
                        {
                            "mode": mode,
                            "producer_id": producer.unique_id,
                            "alpha": round(producer.alpha, 4),
                            "task_type": task.task_type,
                            "complexity_index": task.complexity_index,
                            "investment_step": original,
                        }
                    )
    return pd.DataFrame(
        timing_rows,
        columns=["mode", "producer_id", "alpha", "task_type", "complexity_index", "investment_step"],
    )


def export_bundle(bundle: RunBundle, output_dir: Path | str):
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    labour_figure = build_labour_market_figure(bundle.results)
    labour_path = output_path / "comparison_labour_market.png"
    labour_figure.savefig(labour_path, dpi=150, bbox_inches="tight")
    plt.close(labour_figure)

    macro_figure = build_macro_figure(bundle.results)
    macro_path = output_path / "comparison_macro.png"
    macro_figure.savefig(macro_path, dpi=150, bbox_inches="tight")
    plt.close(macro_figure)

    complexity_figure = build_automation_complexity_figure(bundle)
    complexity_path = output_path / "automation_complexity.png"
    complexity_figure.savefig(complexity_path, dpi=150, bbox_inches="tight")
    plt.close(complexity_figure)

    excel_path = output_path / "simulation_results.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        build_summary_dataframe(bundle).to_excel(writer, sheet_name="Summary")
        for mode in bundle.modes:
            sheet_name = mode.replace("npv_", "").upper()[:31]
            bundle.results[mode].to_excel(writer, sheet_name=sheet_name)
        build_combined_dataframe(bundle).to_excel(writer, sheet_name="Combined")
        build_producer_dataframe(bundle).to_excel(writer, sheet_name="Producer_Tasks", index=False)
        build_timing_dataframe(bundle).to_excel(writer, sheet_name="Adoption_Timing", index=False)
        pd.DataFrame(
            [{"parameter": key, "value": str(value)} for key, value in bundle.params.items()]
            + [{"parameter": "N_STEPS", "value": str(bundle.n_steps)}]
        ).to_excel(writer, sheet_name="Parameters", index=False)

    return {
        "labour_market_png": labour_path,
        "macro_png": macro_path,
        "automation_complexity_png": complexity_path,
        "excel": excel_path,
    }


def _history_root(base_dir: Path | str) -> Path:
    return Path(base_dir) / "dashboard_runs"


def save_run_to_history(bundle: RunBundle, base_dir: Path | str) -> Path:
    history_dir = _history_root(base_dir)
    history_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_label = bundle.run_label or "run"
    safe_label = "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in run_label).strip("_")
    safe_label = safe_label or "run"
    run_id = f"{timestamp}_{safe_label}"
    run_dir = history_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    bundle.run_id = run_id

    combined = build_combined_dataframe(bundle)
    summary = build_summary_dataframe(bundle)
    combined.to_csv(run_dir / "combined.csv")
    summary.to_csv(run_dir / "summary.csv")
    export_bundle(bundle, run_dir)

    metadata = {
        "run_id": run_id,
        "run_label": bundle.run_label or run_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "n_steps": bundle.n_steps,
        "modes": bundle.modes,
        "params": bundle.params,
        "final_metrics": {
            mode: {
                metric: float(bundle.results[mode][metric].iloc[-1])
                for metric in KEY_METRICS
                if metric in bundle.results[mode].columns
            }
            for mode in bundle.modes
        },
    }
    (run_dir / "metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    return run_dir


def load_run_history(base_dir: Path | str) -> list[dict]:
    history_dir = _history_root(base_dir)
    if not history_dir.exists():
        return []

    runs = []
    for run_dir in sorted(history_dir.iterdir(), reverse=True):
        metadata_path = run_dir / "metadata.json"
        combined_path = run_dir / "combined.csv"
        if not metadata_path.exists() or not combined_path.exists():
            continue
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        metadata["path"] = str(run_dir)
        runs.append(metadata)
    return runs


def load_saved_run_bundle(base_dir: Path | str, run_id: str) -> Optional[RunBundle]:
    run_dir = _history_root(base_dir) / run_id
    metadata_path = run_dir / "metadata.json"
    combined_path = run_dir / "combined.csv"
    if not metadata_path.exists() or not combined_path.exists():
        return None

    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    combined = pd.read_csv(combined_path)
    if "step" in combined.columns:
        combined = combined.set_index("step")

    results: Dict[str, pd.DataFrame] = {}
    modes = list(metadata.get("modes", []))
    for mode in modes:
        prefix = f"{mode}__"
        mode_columns = [column for column in combined.columns if column.startswith(prefix)]
        if not mode_columns:
            continue
        mode_frame = combined[mode_columns].copy()
        mode_frame.columns = [column[len(prefix):] for column in mode_columns]
        results[mode] = ensure_derived_metrics(mode_frame)

    return RunBundle(
        results=results,
        models={},
        params=dict(metadata.get("params", {})),
        n_steps=int(metadata.get("n_steps", 0)),
        modes=modes,
        run_label=metadata.get("run_label", run_id),
        run_id=run_id,
    )


def _make_writable_and_retry(func, path, exc_info):
    try:
        os.chmod(path, stat.S_IWRITE)
        func(path)
    except Exception:
        pass


def _remove_tree_robust(path: Path, retries: int = 4, delay_seconds: float = 0.35) -> bool:
    for attempt in range(retries):
        try:
            shutil.rmtree(path, onerror=_make_writable_and_retry)
            return True
        except FileNotFoundError:
            return True
        except PermissionError:
            if attempt == retries - 1:
                return False
            time.sleep(delay_seconds)
        except OSError:
            if attempt == retries - 1:
                return False
            time.sleep(delay_seconds)
    return not path.exists()


def delete_run_from_history(base_dir: Path | str, run_id: str) -> bool:
    run_dir = _history_root(base_dir) / run_id
    if not run_dir.exists() or not run_dir.is_dir():
        return False
    return _remove_tree_robust(run_dir)


def rename_run_in_history(base_dir: Path | str, run_id: str, new_label: str) -> bool:
    run_dir = _history_root(base_dir) / run_id
    metadata_path = run_dir / "metadata.json"
    if not metadata_path.exists():
        return False
    cleaned_label = new_label.strip()
    if not cleaned_label:
        return False
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    metadata["run_label"] = cleaned_label
    metadata_path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    return True


def load_combined_history_frame(base_dir: Path | str, run_ids: Iterable[str], metric: str, modes: Optional[Iterable[str]] = None) -> pd.DataFrame:
    history_dir = _history_root(base_dir)
    selected_modes = list(modes or MODES)
    frames = []
    for run_id in run_ids:
        combined_path = history_dir / run_id / "combined.csv"
        metadata_path = history_dir / run_id / "metadata.json"
        if not combined_path.exists() or not metadata_path.exists():
            continue
        combined = pd.read_csv(combined_path)
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if "step" in combined.columns:
            combined = combined.set_index("step")
        created_at = metadata.get("created_at", "")
        display_label = f"{metadata.get('run_label', run_id)} [{created_at}]" if created_at else metadata.get("run_label", run_id)
        for mode in selected_modes:
            column = f"{mode}__{metric}"
            if column not in combined.columns and metric == "skill_wage_premium":
                high_col = f"{mode}__wage_high"
                low_col = f"{mode}__wage_low"
                if high_col in combined.columns and low_col in combined.columns:
                    wage_high = pd.to_numeric(combined[high_col], errors="coerce")
                    wage_low = pd.to_numeric(combined[low_col], errors="coerce")
                    combined[column] = wage_high.div(wage_low.where(wage_low > 0, np.nan))
            if column not in combined.columns:
                continue
            frame = combined[[column]].rename(columns={column: "value"}).reset_index()
            frame["run_id"] = run_id
            frame["run_label"] = metadata.get("run_label", run_id)
            frame["run_display"] = display_label
            frame["mode"] = mode
            frames.append(frame)
    if not frames:
        return pd.DataFrame(columns=["step", "value", "run_id", "run_label", "run_display", "mode"])
    return pd.concat(frames, ignore_index=True)


def build_comparison_figure(history_frame: pd.DataFrame, metric: str, color_by: str = "run", title_suffix: str = ""):
    fig = plt.figure(figsize=(12, 5))
    grid = gridspec.GridSpec(1, 1)
    ax = fig.add_subplot(grid[0])
    frame = history_frame.sort_values("step").copy()
    series_cols = ["run_id", "run_display", "mode"]
    if metric in CUMULATIVE_CHANNEL_METRICS:
        frame["value"] = frame.groupby(series_cols)["value"].cumsum()
    if color_by == "run":
        color_keys = list(dict.fromkeys(frame["run_id"].tolist()))
        palette = plt.cm.tab10.colors if len(color_keys) <= 10 else plt.cm.tab20.colors
        colors = {key: mcolors.to_hex(palette[index % len(palette)]) for index, key in enumerate(color_keys)}
    else:
        colors = MODE_COLORS

    grouped_items = list(frame.groupby(series_cols, sort=False))
    if metric in CUMULATIVE_CHANNEL_METRICS:
        grouped_items = sorted(
            grouped_items,
            key=lambda item: float(item[1]["value"].iloc[-1]) if not item[1].empty else 0.0,
            reverse=True,
        )
    for draw_order, ((run_id, run_display, mode), subset) in enumerate(grouped_items):
        if color_by == "run":
            color = colors.get(run_id, "#1f77b4")
            linestyle = MODE_LS.get(mode, "-")
            label = f"{run_display} | {MODE_LABELS.get(mode, mode)}"
        else:
            color = colors.get(mode, "#1f77b4")
            linestyle = MODE_LS.get(mode, "-")
            label = MODE_LABELS.get(mode, mode)
        if metric in CUMULATIVE_CHANNEL_METRICS:
            ax.plot(
                subset["step"],
                subset["value"],
                label=label,
                color=color,
                linestyle=linestyle,
                marker=MODE_MARKERS.get(mode, None),
                markersize=3.2,
                markevery=max(1, len(subset.index) // 14),
                linewidth=2.2,
                alpha=0.9,
                zorder=10 + draw_order,
            )
        else:
            ax.plot(
                subset["step"],
                subset["value"],
                label=label,
                color=color,
                linestyle=linestyle,
                linewidth=2.2,
            )
    title = METRIC_LABELS.get(metric, metric)
    if metric in CUMULATIVE_CHANNEL_METRICS:
        title = f"Cumulative {title}"
    if title_suffix:
        title = f"{title} - {title_suffix}"
    ax.set_title(title, fontsize=11, fontweight="bold")
    ax.set_xlabel("Model step", fontsize=9)
    ax.set_ylabel(
        f"Cumulative {METRIC_LABELS.get(metric, metric)}"
        if metric in CUMULATIVE_CHANNEL_METRICS
        else METRIC_LABELS.get(metric, metric),
        fontsize=9,
    )
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=8, loc="best")
    fig.tight_layout()
    return fig


# =============================================================================
# Experimenter - run and compare multiple named configurations
# =============================================================================

def run_experiment_batch(
    configs: list[dict],
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
    log_decisions: bool = False,
) -> list[dict]:
    """
    Run a batch of named experiment configurations.

    Each config must have:
        name     - display label (str)
        params   - full parameter dict (merged with BASE_PARAMS)
        modes    - list of adoption modes to run
        n_steps  - number of simulation steps (int)

    Returns a list of {"name": str, "bundle": RunBundle} dicts.
    """
    results = []
    total_ticks = sum(len(cfg["modes"]) * cfg["n_steps"] for cfg in configs)
    completed = 0
    for cfg in configs:
        name = cfg["name"]
        resolved_params = {**BASE_PARAMS, **cfg["params"]}
        bundle = RunBundle(results={}, models={}, params=resolved_params,
                           n_steps=cfg["n_steps"], modes=cfg["modes"], run_label=name)
        for mode in cfg["modes"]:
            model_params = {**resolved_params, "adoption_mode": mode, "log_decisions": log_decisions}
            model = LabourMarketModel(**model_params)
            for _ in range(cfg["n_steps"]):
                model.step()
                completed += 1
                if progress_callback:
                    progress_callback(f"{name} / {mode}", completed, total_ticks)
            bundle.results[mode] = ensure_derived_metrics(model.datacollector.get_model_vars_dataframe())
            bundle.models[mode] = model
        results.append({"name": name, "bundle": bundle})
    return results


def build_experiment_comparison_figure(
    exp_results: list[dict],
    metric: str,
    mode: str,
    rolling: int = 1,
) -> "plt.Figure":
    """
    One series per experiment (named configuration) for a given metric and mode.
    Colors cycle through tab10/tab20.
    """
    palette = list(plt.cm.tab10.colors) + list(plt.cm.tab20.colors[10:])
    _linestyles = ["-", "--", ":", "-."]
    fig, ax = plt.subplots(figsize=(12, 5))
    n_plotted = 0
    npv_only_note = False
    _use_rolling = rolling > 1
    for idx, item in enumerate(exp_results):
        name = item["name"]
        bundle = item["bundle"]
        df = bundle.results.get(mode)
        if df is not None:
            df = ensure_derived_metrics(df)
        if df is None or metric not in df.columns:
            continue
        series = pd.to_numeric(df[metric], errors="coerce")
        # Skip series that are entirely NaN (e.g. new_automations_this_step in ULC mode)
        if series.isna().all():
            npv_only_note = True
            continue
        color = mcolors.to_hex(palette[idx % len(palette)])
        ls = _linestyles[idx % len(_linestyles)]
        if metric in CUMULATIVE_CHANNEL_METRICS:
            ax.plot(
                df.index,
                series.fillna(0).cumsum(),
                label=name,
                color=color,
                linestyle=ls,
                linewidth=2.2,
                alpha=0.9,
            )
        elif _use_rolling:
            smoothed = series.rolling(rolling, min_periods=1, center=True).mean()
            ax.plot(df.index, series, color=color, linewidth=0.8, alpha=0.20, linestyle=ls)
            ax.plot(df.index, smoothed, label=name, color=color, linewidth=2.2, linestyle=ls)
        else:
            ax.plot(df.index, series, label=name, color=color, linewidth=1.6, linestyle=ls)
        n_plotted += 1
    label = METRIC_LABELS.get(metric, metric)
    mode_label = MODE_LABELS.get(mode, mode)
    rolling_suffix = f"  ({rolling}-step avg)" if _use_rolling else ""
    if metric in CUMULATIVE_CHANNEL_METRICS:
        ax.set_title(f"Cumulative {label} - {mode_label}", fontsize=11, fontweight="bold")
        ax.set_ylabel(f"Cumulative {label}", fontsize=9)
    else:
        ax.set_title(f"{label} - {mode_label}{rolling_suffix}", fontsize=11, fontweight="bold")
        ax.set_ylabel(label, fontsize=9)
    if n_plotted == 0:
        ax.text(0.5, 0.5, "No data for this metric in this mode\n(metric may be NPV-modes only)",
                ha="center", va="center", transform=ax.transAxes, fontsize=10, color="#94a3b8")
    elif npv_only_note:
        ax.set_title(
            ax.get_title() + "  ⚠ some experiments had no data (NPV modes only)",
            fontsize=10, fontweight="bold",
        )
    ax.set_xlabel("Model step", fontsize=9)
    ax.grid(True, linestyle="--", alpha=0.5)
    if metric in ("ai_adoption_rate", "employment_rate_high", "employment_rate_low", "labour_share"):
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    if n_plotted > 0:
        legend_title = f"Bold = {rolling}-step rolling avg  |  faint = raw" if _use_rolling else None
        ax.legend(fontsize=8, loc="best", title=legend_title, title_fontsize=7)
    fig.tight_layout()
    return fig


def build_experiment_final_table(exp_results: list[dict], mode: str) -> pd.DataFrame:
    """
    Summary table: one row per experiment, columns are final-step key metrics.
    """
    rows = []
    for item in exp_results:
        name = item["name"]
        bundle = item["bundle"]
        df = bundle.results.get(mode)
        if df is not None:
            df = ensure_derived_metrics(df)
        row = {"experiment": name}
        for metric in KEY_METRICS:
            if df is not None and metric in df.columns:
                row[METRIC_LABELS.get(metric, metric)] = df[metric].iloc[-1]
            else:
                row[METRIC_LABELS.get(metric, metric)] = float("nan")
        rows.append(row)
    return pd.DataFrame(rows).set_index("experiment")


# =============================================================================
# Sensitivity sweep - vary one parameter across a range
# =============================================================================

def run_sensitivity_sweep(
    base_params: Dict,
    param_name: str,
    param_values: list,
    n_steps: int,
    mode: str,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
) -> list[dict]:
    """
    Run one simulation per value in param_values, varying only param_name.
    Returns a list of dicts:
        {"param_value": v, "df": DataFrame of time-series, "final": dict of metric->value}
    """
    results = []
    total_ticks = len(param_values) * n_steps
    completed = 0
    for v in param_values:
        run_params = {**BASE_PARAMS, **base_params, param_name: v, "adoption_mode": mode}
        model = LabourMarketModel(**run_params)
        for _ in range(n_steps):
            model.step()
            completed += 1
            if progress_callback:
                progress_callback(f"{param_name}={v}", completed, total_ticks)
        df = ensure_derived_metrics(model.datacollector.get_model_vars_dataframe())
        final = {m: float(df[m].iloc[-1]) for m in KEY_METRICS if m in df.columns}
        results.append({"param_value": v, "df": df, "final": final, "model": model})
    return results


def build_sensitivity_timeseries_figure(
    sweep_results: list[dict],
    param_name: str,
    metric: str,
) -> "plt.Figure":
    """
    Time-series for a single metric across all parameter values.
    Lines are colored on a continuous gradient from the colormap.
    """
    n = len(sweep_results)
    cmap = plt.cm.viridis
    fig, ax = plt.subplots(figsize=(12, 5))
    for idx, item in enumerate(sweep_results):
        v = item["param_value"]
        df = item["df"]
        if metric not in df.columns:
            continue
        color = cmap(idx / max(n - 1, 1))
        ax.plot(df.index, df[metric], label=f"{param_name}={v}", color=color, linewidth=1.6)
    label = METRIC_LABELS.get(metric, metric)
    ax.set_title(f"{label} - varying {param_name}", fontsize=11, fontweight="bold")
    ax.set_xlabel("Model step", fontsize=9)
    ax.set_ylabel(label, fontsize=9)
    if metric in ("ai_adoption_rate", "employment_rate_high", "employment_rate_low", "labour_share"):
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=7, loc="best", ncol=max(1, n // 8))
    # Color bar to show gradient
    sm = plt.cm.ScalarMappable(
        cmap=cmap,
        norm=plt.Normalize(
            vmin=sweep_results[0]["param_value"],
            vmax=sweep_results[-1]["param_value"],
        ),
    )
    sm.set_array([])
    fig.colorbar(sm, ax=ax, label=param_name, fraction=0.025, pad=0.01)
    fig.tight_layout()
    return fig


def build_sensitivity_final_figure(
    sweep_results: list[dict],
    param_name: str,
    metrics: list[str],
) -> "plt.Figure":
    """
    Multi-panel plot: for each metric in `metrics`, one subplot showing
    final-step value vs param_value.
    """
    n = len(metrics)
    ncols = min(3, n)
    nrows = (n + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(6 * ncols, 4 * nrows))
    axes_flat = np.array(axes).flatten() if n > 1 else [axes]

    param_vals = [item["param_value"] for item in sweep_results]

    for ax, metric in zip(axes_flat, metrics):
        final_vals = [item["final"].get(metric, float("nan")) for item in sweep_results]
        label = METRIC_LABELS.get(metric, metric)
        ax.plot(param_vals, final_vals, marker="o", linewidth=2.0, color="#2563eb")
        ax.set_title(label, fontsize=9, fontweight="bold")
        ax.set_xlabel(param_name, fontsize=8)
        ax.set_ylabel(label, fontsize=8)
        if metric in ("ai_adoption_rate", "employment_rate_high", "employment_rate_low", "labour_share"):
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
        ax.tick_params(labelsize=7)
        ax.grid(True, linestyle="--", alpha=0.5)

    for ax in axes_flat[n:]:
        ax.set_visible(False)

    fig.suptitle(f"Final-step metrics vs {param_name}", fontsize=12, fontweight="bold")
    fig.tight_layout()
    return fig


def build_sensitivity_summary_table(sweep_results: list[dict], param_name: str) -> pd.DataFrame:
    rows = []
    for item in sweep_results:
        # Single-param sweep: param_value is a scalar; multi-param: param_values is a dict
        if "param_values" in item:
            row = {k: v for k, v in item["param_values"].items()}
        else:
            row = {param_name: item["param_value"]}
        row.update({METRIC_LABELS.get(m, m): v for m, v in item["final"].items()})
        rows.append(row)
    df = pd.DataFrame(rows)
    index_cols = list(item["param_values"].keys()) if "param_values" in sweep_results[0] else [param_name]
    return df.set_index(index_cols) if len(index_cols) == 1 else df.set_index(index_cols[0])


def build_sensitivity_multirun_timeseries(
    sweep_results: list[dict],
    metric: str,
    step_min: int = 0,
    step_max: Optional[int] = None,
    highlight_indices: Optional[list[int]] = None,
    rolling: int = 1,
) -> "plt.Figure":
    """
    Time-series for a single metric across all runs in a sweep (single or multi-param).
    Each run is one line. No legend is shown (too cluttered for many runs).
    A colorbar is shown for single-param sweeps when no filter is active.
    When highlight_indices is provided, only those runs are drawn in colour;
    the rest are shown in light grey at low opacity.
    """
    n = len(sweep_results)
    is_multi = "param_values" in sweep_results[0]
    palette = list(plt.cm.tab10.colors) + list(plt.cm.tab20.colors[10:])
    has_filter = highlight_indices is not None
    highlight_set = set(highlight_indices) if has_filter else None
    # Map highlight idx → sequential colour index for stable palette assignment
    hi_colour_map: dict[int, int] = (
        {run_idx: ci for ci, run_idx in enumerate(sorted(highlight_set))}
        if highlight_set else {}
    )

    if not is_multi:
        cmap = plt.cm.viridis
        param_vals = [item["param_value"] for item in sweep_results]
        vmin, vmax = min(param_vals), max(param_vals)

    fig, ax = plt.subplots(figsize=(12, 5))
    label = METRIC_LABELS.get(metric, metric)

    for idx, item in enumerate(sweep_results):
        df = item["df"]
        if metric not in df.columns:
            continue
        series = pd.to_numeric(df[metric], errors="coerce")
        if metric in CUMULATIVE_CHANNEL_METRICS:
            series = series.fillna(0).cumsum()

        # Timestep filter
        if step_max is not None:
            mask = (df.index >= step_min) & (df.index <= step_max)
        else:
            mask = df.index >= step_min
        x = df.index[mask]
        y = series[mask]

        # Determine colour and visibility
        if has_filter and idx not in highlight_set:
            ax.plot(x, y, color="#cccccc", linewidth=0.8, alpha=0.35, zorder=1)
            continue

        if is_multi:
            ci = hi_colour_map[idx] if has_filter else idx
            color = palette[ci % len(palette)]
        else:
            color = cmap((param_vals[idx] - vmin) / max(vmax - vmin, 1e-9))

        lw = 2.0 if has_filter else 1.6
        _use_rolling = rolling > 1 and metric not in CUMULATIVE_CHANNEL_METRICS
        if _use_rolling:
            smoothed = y.rolling(rolling, min_periods=1, center=True).mean()
            ax.plot(x, y, color=color, linewidth=0.6, alpha=0.20, zorder=2)
            ax.plot(x, smoothed, color=color, linewidth=lw, alpha=0.9, zorder=3)
        else:
            ax.plot(x, y, color=color, linewidth=lw, alpha=0.9, zorder=2)

    _use_rolling_title = rolling > 1 and metric not in CUMULATIVE_CHANNEL_METRICS
    rolling_suffix = f"  ({rolling}-step avg)" if _use_rolling_title else ""
    if metric in CUMULATIVE_CHANNEL_METRICS:
        ax.set_title(f"Cumulative {label} — sensitivity sweep", fontsize=11, fontweight="bold")
        ax.set_ylabel(f"Cumulative {label}", fontsize=9)
    else:
        ax.set_title(f"{label} — sensitivity sweep{rolling_suffix}", fontsize=11, fontweight="bold")
        ax.set_ylabel(label, fontsize=9)

    ax.set_xlabel("Model step", fontsize=9)
    ax.grid(True, linestyle="--", alpha=0.5)
    if metric in ("ai_adoption_rate", "employment_rate_high", "employment_rate_low", "labour_share"):
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))

    # Colorbar for single-param sweeps (no filter active, otherwise colour meaning changes)
    if not is_multi and n > 1 and not has_filter:
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(vmin=vmin, vmax=vmax))
        sm.set_array([])
        param_name = sweep_results[0].get("param_name", "param")
        fig.colorbar(sm, ax=ax, label=param_name, fraction=0.025, pad=0.01)

    if has_filter:
        n_match = len(highlight_set)
        ax.set_title(
            f"{label} — {n_match}/{n} runs match filter",
            fontsize=11, fontweight="bold",
        )

    fig.tight_layout()
    return fig


# ---- LHS / Monte Carlo / Factorial sampling --------------------------------

def _lhs_samples(param_ranges: list[tuple], n_samples: int, seed: int = 0) -> np.ndarray:
    """
    Latin Hypercube Sampling: returns (n_samples, n_params) array in [0,1].
    param_ranges is unused here (just determines n_params); caller maps to actual values.
    """
    rng = np.random.default_rng(seed)
    n_params = len(param_ranges)
    result = np.zeros((n_samples, n_params))
    for j in range(n_params):
        perm = rng.permutation(n_samples)
        result[:, j] = (perm + rng.random(n_samples)) / n_samples
    return result


def _generate_sample_matrix(
    param_configs: list[dict],  # [{name, min, max, is_int}]
    method: str,                # "lhs" | "monte_carlo" | "factorial"
    n_samples: int,             # for lhs/mc; per-param grid size for factorial
    seed: int = 0,
) -> list[dict]:
    """
    Returns a list of param-value dicts, one per run.
    """
    rng = np.random.default_rng(seed)
    n_params = len(param_configs)

    if method == "factorial":
        # Full factorial: n_samples grid points per param
        grids = []
        for pc in param_configs:
            pts = np.linspace(pc["min"], pc["max"], n_samples)
            if pc.get("is_int"):
                pts = np.unique(pts.round().astype(int))
            grids.append(pts)
        combos = np.array(np.meshgrid(*grids, indexing="ij")).reshape(n_params, -1).T
        return [
            {pc["name"]: (int(round(v)) if pc.get("is_int") else float(v))
             for pc, v in zip(param_configs, row)}
            for row in combos
        ]

    elif method == "lhs":
        unit = _lhs_samples(param_configs, n_samples, seed)
    else:  # monte_carlo
        unit = rng.random((n_samples, n_params))

    rows = []
    for row in unit:
        d = {}
        for j, pc in enumerate(param_configs):
            v = pc["min"] + row[j] * (pc["max"] - pc["min"])
            d[pc["name"]] = int(round(v)) if pc.get("is_int") else float(v)
        rows.append(d)
    return rows


def run_sensitivity_sweep_multi(
    base_params: dict,
    param_configs: list[dict],   # [{name, min, max, is_int}]
    method: str,
    n_samples: int,
    n_steps: int,
    mode: str,
    seed: int = 0,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
) -> list[dict]:
    """
    Multi-parameter sensitivity sweep.
    Returns list of {"param_values": {p:v,...}, "df": DataFrame, "final": {metric: v}}.
    """
    sample_matrix = _generate_sample_matrix(param_configs, method, n_samples, seed)
    results = []
    total_ticks = len(sample_matrix) * n_steps
    completed = 0
    for run_idx, pv_dict in enumerate(sample_matrix):
        run_params = {**BASE_PARAMS, **base_params, **pv_dict, "adoption_mode": mode}
        model = LabourMarketModel(**run_params)
        label = ", ".join(f"{k}={v}" for k, v in pv_dict.items())
        for _ in range(n_steps):
            model.step()
            completed += 1
            if progress_callback:
                progress_callback(f"Run {run_idx+1}/{len(sample_matrix)}: {label}", completed, total_ticks)
        df = ensure_derived_metrics(model.datacollector.get_model_vars_dataframe())
        final = {m: float(df[m].iloc[-1]) for m in KEY_METRICS if m in df.columns}
        results.append({"param_values": pv_dict, "df": df, "final": final, "model": model})
    return results


def make_single_run_bundle(item: dict, mode: str) -> "RunBundle":
    """
    Wrap a single sensitivity-sweep result dict into a minimal RunBundle so that
    workforce figures (which expect a bundle with .models) can be used on it.
    The item must contain a 'model' key (available for live sweep results;
    not present in loaded sweeps).
    """
    model = item.get("model")
    if model is None:
        return RunBundle(results={}, models={}, params={}, n_steps=0, modes=[mode])
    return RunBundle(
        results={mode: item["df"]},
        models={mode: model},
        params={},
        n_steps=len(item["df"]),
        modes=[mode],
        run_label="sensitivity_run",
    )


# ---- Sensitivity setup save / load ----------------------------------------

def _sens_setup_root(base_dir: Path | str) -> Path:
    return Path(base_dir) / "dashboard_sensitivity_setups"


def save_sensitivity_setup(setup: dict, base_dir: Path | str, name: str) -> Path:
    """
    Persist a sensitivity setup (params, ranges, method) as JSON so it can be reloaded.
    """
    root = _sens_setup_root(base_dir)
    root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in name).strip("_") or "setup"
    path = root / f"{safe}_{timestamp}.json"
    path.write_text(json.dumps({**setup, "name": name, "created_at": datetime.now().isoformat(timespec="seconds")}, indent=2, default=str), encoding="utf-8")
    return path


def list_sensitivity_setups(base_dir: Path | str) -> list[dict]:
    root = _sens_setup_root(base_dir)
    if not root.exists():
        return []
    setups = []
    for p in sorted(root.glob("*.json"), reverse=True):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            d["path"] = str(p)
            setups.append(d)
        except Exception:
            pass
    return setups


# =============================================================================
# Experiment fingerprinting - cache key for incremental runs
# =============================================================================

def _config_fingerprint(cfg: dict) -> str:
    """Stable hash of an experiment config for change detection."""
    stable = json.dumps(
        {
            "name": cfg["name"],
            "params": {k: str(v) for k, v in sorted(cfg["params"].items())},
            "modes": sorted(cfg["modes"]),
            "n_steps": cfg["n_steps"],
        },
        sort_keys=True,
    )
    return hashlib.sha1(stable.encode()).hexdigest()[:12]


# =============================================================================
# Save / load - experimenter batches
# =============================================================================

def _exp_root(base_dir: Path | str) -> Path:
    return Path(base_dir) / "dashboard_experimenter"


def save_experiment_batch(
    exp_results: list[dict],
    configs: list[dict],
    base_dir: Path | str,
    batch_name: str,
) -> Path:
    """
    Persist a named experiment batch so it can be loaded in a future session.

    Layout: dashboard_experimenter/<safe_name>_<timestamp>/
        metadata.json   - batch name, config list, fingerprints, created_at
        <exp_name>_<mode>.csv  - one CSV per (experiment, mode)
    """
    root = _exp_root(base_dir)
    root.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in batch_name).strip("_") or "batch"
    batch_dir = root / f"{safe}_{timestamp}"
    batch_dir.mkdir(parents=True, exist_ok=True)

    fingerprints = {}
    for cfg, item in zip(configs, exp_results):
        fp = _config_fingerprint(cfg)
        fingerprints[cfg["name"]] = fp
        bundle = item["bundle"]
        for mode, df in bundle.results.items():
            safe_exp = "".join(c if c.isalnum() or c in "-_" else "_" for c in cfg["name"]).strip("_")
            csv_path = batch_dir / f"{safe_exp}__{mode}.csv"
            df.to_csv(csv_path)

    metadata = {
        "batch_name": batch_name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "configs": configs,
        "fingerprints": fingerprints,
    }
    (batch_dir / "metadata.json").write_text(json.dumps(metadata, indent=2, default=str), encoding="utf-8")
    return batch_dir


def list_experiment_batches(base_dir: Path | str) -> list[dict]:
    root = _exp_root(base_dir)
    if not root.exists():
        return []
    batches = []
    for batch_dir in sorted(root.iterdir(), reverse=True):
        meta_path = batch_dir / "metadata.json"
        if not meta_path.exists():
            continue
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        meta["path"] = str(batch_dir)
        batches.append(meta)
    return batches


def load_experiment_batch(batch_path: Path | str) -> tuple[list[dict], list[dict]]:
    """
    Returns (configs, exp_results) where exp_results is the same structure
    as returned by run_experiment_batch.
    """
    batch_dir = Path(batch_path)
    meta = json.loads((batch_dir / "metadata.json").read_text(encoding="utf-8"))
    configs = meta["configs"]
    exp_results = []
    for cfg in configs:
        safe_exp = "".join(c if c.isalnum() or c in "-_" else "_" for c in cfg["name"]).strip("_")
        bundle_results = {}
        for mode in cfg["modes"]:
            csv_path = batch_dir / f"{safe_exp}__{mode}.csv"
            if not csv_path.exists():
                continue
            df = pd.read_csv(csv_path)
            if "Step" in df.columns:
                df = df.set_index("Step")
            elif df.columns[0] not in KEY_METRICS:
                df = df.set_index(df.columns[0])
            bundle_results[mode] = ensure_derived_metrics(df)
        bundle = RunBundle(
            results=bundle_results,
            models={},
            params=cfg["params"],
            n_steps=cfg["n_steps"],
            modes=cfg["modes"],
            run_label=cfg["name"],
        )
        exp_results.append({"name": cfg["name"], "bundle": bundle})
    return configs, exp_results


def delete_experiment_batch(batch_path: Path | str) -> bool:
    batch_dir = Path(batch_path)
    if not batch_dir.exists():
        return False
    shutil.rmtree(batch_dir)
    return True


# =============================================================================
# Save / load - sensitivity sweeps
# =============================================================================

def _sens_root(base_dir: Path | str) -> Path:
    return Path(base_dir) / "dashboard_sensitivity"


def save_sensitivity_sweep(
    sens_state: dict,
    base_dir: Path | str,
    sweep_name: str,
) -> Path:
    """
    Persist a sensitivity sweep (single- or multi-param).

    Layout: dashboard_sensitivity/<safe_name>_<timestamp>/
        metadata.json  - all state except dataframes
        sweep_<i>.csv  - time-series for run i
    """
    root = _sens_root(base_dir)
    root.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in sweep_name).strip("_") or "sweep"
    sweep_dir = root / f"{safe}_{timestamp}"
    sweep_dir.mkdir(parents=True, exist_ok=True)

    sweep = sens_state["sweep"]
    is_multi = "param_values" in sweep[0] if sweep else False
    for idx, item in enumerate(sweep):
        (sweep_dir / f"sweep_{idx:03d}.csv").write_text(
            item["df"].to_csv(), encoding="utf-8"
        )

    metadata = {
        "sweep_name": sweep_name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "sweep_type": "multi" if is_multi else "single",
        "param_name": sens_state.get("param_name", ""),
        "param_values": sens_state.get("param_values", []),
        "param_configs": sens_state.get("param_configs", []),
        "method": sens_state.get("method", "single"),
        "mode": sens_state["mode"],
        "base_params": sens_state.get("base_params", {}),
        "finals": [item["final"] for item in sweep],
        # For multi-param: store each run's param dict
        "run_param_values": [item.get("param_values", {}) for item in sweep] if is_multi else [],
        "single_param_values": [item.get("param_value") for item in sweep] if not is_multi else [],
    }
    (sweep_dir / "metadata.json").write_text(json.dumps(metadata, indent=2, default=str), encoding="utf-8")
    return sweep_dir


def list_sensitivity_sweeps(base_dir: Path | str) -> list[dict]:
    root = _sens_root(base_dir)
    if not root.exists():
        return []
    sweeps = []
    for sweep_dir in sorted(root.iterdir(), reverse=True):
        meta_path = sweep_dir / "metadata.json"
        if not meta_path.exists():
            continue
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        meta["path"] = str(sweep_dir)
        sweeps.append(meta)
    return sweeps


def load_sensitivity_sweep(sweep_path: Path | str) -> dict:
    """Returns a sens_state dict identical to what the run produces (single or multi-param)."""
    sweep_dir = Path(sweep_path)
    meta = json.loads((sweep_dir / "metadata.json").read_text(encoding="utf-8"))
    finals = meta["finals"]
    is_multi = meta.get("sweep_type") == "multi"
    sweep = []
    n_runs = len(finals)
    for idx in range(n_runs):
        csv_path = sweep_dir / f"sweep_{idx:03d}.csv"
        if csv_path.exists():
            df = pd.read_csv(csv_path)
            if df.columns[0] not in KEY_METRICS:
                df = df.set_index(df.columns[0])
        else:
            df = pd.DataFrame()
        df = ensure_derived_metrics(df)
        final = dict(finals[idx])
        if "skill_wage_premium" not in final and "skill_wage_premium" in df.columns and not df.empty:
            final["skill_wage_premium"] = float(df["skill_wage_premium"].iloc[-1])
        if is_multi:
            sweep.append({"param_values": meta["run_param_values"][idx], "df": df, "final": final})
        else:
            sweep.append({"param_value": meta["single_param_values"][idx], "df": df, "final": final})
    return {
        "param_name": meta.get("param_name", ""),
        "param_values": meta.get("param_values", []),
        "param_configs": meta.get("param_configs", []),
        "method": meta.get("method", "single"),
        "mode": meta["mode"],
        "base_params": meta.get("base_params", {}),
        "sweep": sweep,
        "sweep_type": meta.get("sweep_type", "single"),
    }


def delete_sensitivity_sweep(sweep_path: Path | str) -> bool:
    sweep_dir = Path(sweep_path)
    if not sweep_dir.exists():
        return False
    shutil.rmtree(sweep_dir)
    return True


# =============================================================================
# Save / load - OFAT runs
# =============================================================================

def _ofat_root(base_dir: Path | str) -> Path:
    return Path(base_dir) / "ofat_runs"


def list_ofat_batches(base_dir: Path | str) -> list[dict]:
    root = _ofat_root(base_dir)
    if not root.exists():
        return []
    batches = []
    for batch_dir in sorted(root.iterdir(), reverse=True):
        meta_path = batch_dir / "metadata.json"
        summary_path = batch_dir / "summary.csv"
        aggregated_path = batch_dir / "aggregated.csv"
        if not meta_path.exists() or not summary_path.exists() or not aggregated_path.exists():
            continue
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        meta["path"] = str(batch_dir)
        batches.append(meta)
    return batches


def load_ofat_batch(batch_path: Path | str) -> dict:
    batch_dir = Path(batch_path)
    metadata = json.loads((batch_dir / "metadata.json").read_text(encoding="utf-8"))
    summary = pd.read_csv(batch_dir / "summary.csv")
    aggregated = pd.read_csv(batch_dir / "aggregated.csv")
    summary = _repair_ofat_frame(summary, metadata)
    aggregated = _repair_ofat_frame(aggregated, metadata)
    return {
        "metadata": metadata,
        "summary": summary,
        "aggregated": aggregated,
        "path": str(batch_dir),
        "timeseries_dir": str(batch_dir / "timeseries"),
    }


def delete_ofat_batch(batch_path: Path | str) -> bool:
    batch_dir = Path(batch_path)
    if not batch_dir.exists():
        return False
    return _remove_tree_robust(batch_dir)


def _repair_ofat_frame(df: pd.DataFrame, metadata: dict) -> pd.DataFrame:
    if df.empty or "parameter" not in df.columns or "grid_index" not in df.columns:
        return df

    repaired = df.copy()
    parameter_grids = metadata.get("parameter_grids", {})
    default_params = metadata.get("dashboard_default_params", {})

    grid_lookup: dict[tuple[str, int], dict] = {}
    for parameter, grid in parameter_grids.items():
        for item in grid:
            try:
                grid_lookup[(parameter, int(item["grid_index"]))] = item
            except Exception:
                continue

    param_values = []
    default_values = []
    factor_labels = []
    factor_pcts = []

    for _, row in repaired.iterrows():
        parameter = row.get("parameter")
        try:
            grid_index = int(row.get("grid_index"))
        except Exception:
            grid_index = None

        entry = grid_lookup.get((parameter, grid_index)) if grid_index is not None else None
        if entry is not None:
            param_values.append(entry.get("param_value"))
            factor_labels.append(entry.get("factor_label", row.get("factor_label")))
            factor_pcts.append(entry.get("factor_pct", row.get("factor_pct")))
        else:
            raw_value = row.get("param_value")
            if isinstance(raw_value, str):
                if raw_value == "True":
                    raw_value = True
                elif raw_value == "False":
                    raw_value = False
                else:
                    maybe_numeric = pd.to_numeric(pd.Series([raw_value]), errors="coerce").iloc[0]
                    raw_value = raw_value if pd.isna(maybe_numeric) else maybe_numeric
            param_values.append(raw_value)
            factor_labels.append(row.get("factor_label"))
            factor_pcts.append(row.get("factor_pct"))

        default_values.append(default_params.get(parameter, row.get("default_value")))

    repaired["param_value"] = param_values
    repaired["default_value"] = default_values
    if "factor_label" in repaired.columns:
        repaired["factor_label"] = factor_labels
    if "factor_pct" in repaired.columns:
        repaired["factor_pct"] = factor_pcts
    return repaired


def build_productivity_preview(alpha: float, params: Dict[str, float], n_tasks: int = 20) -> pd.DataFrame:
    n_routine = round(alpha * n_tasks)
    n_nonroutine = n_tasks - n_routine
    rows = []

    for task_id in range(1, n_tasks + 1):
        if task_id <= n_routine:
            task_type = "routine"
            complexity_index = task_id
        else:
            task_type = "non_routine"
            complexity_index = task_id - n_routine

        productivity_low = _preview_productivity_human(
            params, task_type, complexity_index, "low"
        )
        productivity_high = _preview_productivity_human(
            params, task_type, complexity_index, "high"
        )
        productivity_ai = _preview_productivity_ai(params, task_type, complexity_index)

        cheapest = max(
            [
                ("low", productivity_low),
                ("high", productivity_high),
                ("AI", productivity_ai),
            ],
            key=lambda item: item[1],
        )[0]

        rows.append(
            {
                "task_id": task_id,
                "task_type": task_type,
                "complexity_index": complexity_index,
                "productivity_low": round(float(productivity_low), 4),
                "productivity_high": round(float(productivity_high), 4),
                "productivity_ai": round(float(productivity_ai), 4),
                "highest_productivity": cheapest,
            }
        )

    return pd.DataFrame(rows)


def build_productivity_preview_figure(productivity_df: pd.DataFrame):
    routine_count = int((productivity_df["task_type"] == "routine").sum())
    x = np.arange(len(productivity_df))
    width = 0.25

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(x - width, productivity_df["productivity_low"], width=width, label="Low-skill productivity", color="#f97316")
    ax.bar(x, productivity_df["productivity_high"], width=width, label="High-skill productivity", color="#2563eb")
    ax.bar(x + width, productivity_df["productivity_ai"], width=width, label="AI productivity", color="#16a34a")

    if 0 < routine_count < len(productivity_df):
        ax.axvline(routine_count - 0.5, color="#475569", linestyle=":", linewidth=2.0, label="Routine / non-routine split")

    ax.set_title("Task-by-task productivity preview", fontsize=11, fontweight="bold")
    ax.set_xlabel("Task ID", fontsize=9)
    ax.set_ylabel("Productivity", fontsize=9)
    ax.set_xticks(x)
    ax.set_xticklabels(productivity_df["task_id"].tolist())
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=8, loc="best")
    fig.tight_layout()
    return fig


def build_wage_preview_figure(params: Dict[str, float], n_high_skilled: int, n_low_skilled: int):
    max_workers = max(1, n_high_skilled + n_low_skilled)
    worker_range = np.arange(0, max_workers + 1)
    wage_high = params["a_h"] + params["b_h"] * worker_range
    wage_low = params["a_l"] + params["b_l"] * worker_range
    high_y = params["a_h"] + params["b_h"] * n_high_skilled
    low_y = params["a_l"] + params["b_l"] * n_low_skilled

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(worker_range, wage_high, label="High-skill wage target", color="#2563eb", linewidth=2.2)
    ax.plot(worker_range, wage_low, label="Low-skill wage target", color="#f97316", linewidth=2.2)
    w_min = float(params.get("w_min", BASE_PARAMS.get("w_min", 3.5)))
    if w_min > 0:
        ax.axhline(w_min, label=f"Minimum wage (w_min = {w_min:.2f})", color="#dc2626", linestyle="-.", linewidth=2.0)
    ax.axhline(params["k_ai"], label="AI cost start", color="#111827", linestyle="--", linewidth=1.8)
    ax.axhline(params["k_ai_floor"], label="AI cost floor", color="#16a34a", linestyle=":", linewidth=2.2)
    ax.axvline(n_high_skilled, label=f"n_high_skilled = {n_high_skilled}", color="#2563eb", linestyle="--", linewidth=1.4, alpha=0.8)
    ax.axvline(n_low_skilled, label=f"n_low_skilled = {n_low_skilled}", color="#f97316", linestyle="--", linewidth=1.4, alpha=0.8)
    ax.scatter([n_high_skilled], [high_y], color="#2563eb", s=45, zorder=5)
    ax.scatter([n_low_skilled], [low_y], color="#f97316", s=45, zorder=5)
    ax.annotate(
        f"({n_high_skilled}, {high_y:.2f})",
        xy=(n_high_skilled, high_y),
        xytext=(8, 8),
        textcoords="offset points",
        color="#2563eb",
        fontsize=8,
        bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="#2563eb", alpha=0.85),
    )
    ax.annotate(
        f"({n_low_skilled}, {low_y:.2f})",
        xy=(n_low_skilled, low_y),
        xytext=(8, -18),
        textcoords="offset points",
        color="#c2410c",
        fontsize=8,
        bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="#f97316", alpha=0.85),
    )
    ax.set_title("Wage curves and AI cost reference", fontsize=11, fontweight="bold")
    ax.set_xlabel("Number of employed workers", fontsize=9)
    ax.set_ylabel("€ (wage / AI cost)", fontsize=9)
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=8, loc="best")
    fig.tight_layout()
    return fig


def estimate_initial_assignment(alpha: float, params: Dict[str, float], n_tasks_per_producer: int = 20) -> Dict[str, float]:
    n_producers = int(params["n_producers"])
    total_tasks = n_producers * n_tasks_per_producer
    available_high = int(params["n_high_skilled"])
    available_low = int(params["n_low_skilled"])

    preview_df = build_productivity_preview(alpha, params, n_tasks=n_tasks_per_producer)
    wage_high = params["a_h"] + params["b_h"] * available_high
    wage_low = params["a_l"] + params["b_l"] * available_low

    preferred_high = 0
    preferred_low = 0
    for _, row in preview_df.iterrows():
        uc_high = wage_high / max(row["productivity_high"], 1e-9)
        uc_low = wage_low / max(row["productivity_low"], 1e-9)
        if uc_high <= uc_low:
            preferred_high += 1
        else:
            preferred_low += 1

    preferred_high_total = preferred_high * n_producers
    preferred_low_total = preferred_low * n_producers

    employed_high = min(available_high, preferred_high_total)
    employed_low = min(available_low, preferred_low_total)
    filled_tasks = employed_high + employed_low
    remaining_tasks = max(0, total_tasks - filled_tasks)

    spare_high = max(0, available_high - employed_high)
    spare_low = max(0, available_low - employed_low)

    add_high = min(spare_high, remaining_tasks)
    remaining_tasks -= add_high
    add_low = min(spare_low, remaining_tasks)

    employed_high += add_high
    employed_low += add_low
    filled_tasks = employed_high + employed_low

    return {
        "total_tasks": total_tasks,
        "employed_high": employed_high,
        "employed_low": employed_low,
        "filled_tasks": filled_tasks,
        "unfilled_tasks": max(0, total_tasks - filled_tasks),
        "wage_high": params["a_h"] + params["b_h"] * employed_high,
        "wage_low": params["a_l"] + params["b_l"] * employed_low,
    }


def build_cost_comparison_preview(
    alpha: float,
    params: Dict[str, float],
    n_tasks: int = 20,
    wage_high_override: Optional[float] = None,
    wage_low_override: Optional[float] = None,
    ai_cost_override: Optional[float] = None,
) -> pd.DataFrame:
    productivity_df = build_productivity_preview(alpha, params, n_tasks=n_tasks)
    assignment = estimate_initial_assignment(alpha, params, n_tasks_per_producer=n_tasks)
    wage_high = assignment["wage_high"] if wage_high_override is None else wage_high_override
    wage_low = assignment["wage_low"] if wage_low_override is None else wage_low_override
    ai_cost = params["k_ai"] if ai_cost_override is None else ai_cost_override

    rows = []
    for _, row in productivity_df.iterrows():
        cost_low = wage_low / max(row["productivity_low"], 1e-9)
        cost_high = wage_high / max(row["productivity_high"], 1e-9)
        cost_ai = ai_cost / max(row["productivity_ai"], 1e-9)
        productivity_per_euro_low = row["productivity_low"] / max(wage_low, 1e-9)
        productivity_per_euro_high = row["productivity_high"] / max(wage_high, 1e-9)
        productivity_per_euro_ai = row["productivity_ai"] / max(ai_cost, 1e-9)
        cheapest = min(
            [("low", cost_low), ("high", cost_high), ("AI", cost_ai)],
            key=lambda item: item[1],
        )[0]
        best_value = max(
            [
                ("low", productivity_per_euro_low),
                ("high", productivity_per_euro_high),
                ("AI", productivity_per_euro_ai),
            ],
            key=lambda item: item[1],
        )[0]
        rows.append(
            {
                "task_id": row["task_id"],
                "task_type": row["task_type"],
                "complexity_index": row["complexity_index"],
                "productivity_low": round(float(row["productivity_low"]), 4),
                "productivity_high": round(float(row["productivity_high"]), 4),
                "productivity_ai": round(float(row["productivity_ai"]), 4),
                "unit_cost_low": round(float(cost_low), 4),
                "unit_cost_high": round(float(cost_high), 4),
                "unit_cost_ai": round(float(cost_ai), 4),
                "productivity_per_euro_low": round(float(productivity_per_euro_low), 4),
                "productivity_per_euro_high": round(float(productivity_per_euro_high), 4),
                "productivity_per_euro_ai": round(float(productivity_per_euro_ai), 4),
                "cheapest_input": cheapest,
                "best_value_input": best_value,
            }
        )
    return pd.DataFrame(rows)


def build_cost_comparison_figure(cost_df: pd.DataFrame, metric: str = "unit_cost"):
    routine_count = int((cost_df["task_type"] == "routine").sum())
    x = np.arange(len(cost_df))
    width = 0.25
    if metric == "productivity_per_euro":
        low_column = "productivity_per_euro_low"
        high_column = "productivity_per_euro_high"
        ai_column = "productivity_per_euro_ai"
        ylabel = "Output per € (productivity / wage)"
        title = "Value for money by task"
        low_label = "Low-skill output per €"
        high_label = "High-skill output per €"
        ai_label = "AI output per €"
    else:
        low_column = "unit_cost_low"
        high_column = "unit_cost_high"
        ai_column = "unit_cost_ai"
        ylabel = "€ per unit output (wage / productivity)"
        title = "Cost comparison by task"
        low_label = "Low-skill unit cost (€)"
        high_label = "High-skill unit cost (€)"
        ai_label = "AI unit cost (€)"
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(x - width, cost_df[low_column], width=width, label=low_label, color="#f97316")
    ax.bar(x, cost_df[high_column], width=width, label=high_label, color="#2563eb")
    ax.bar(x + width, cost_df[ai_column], width=width, label=ai_label, color="#16a34a")
    if 0 < routine_count < len(cost_df):
        ax.axvline(routine_count - 0.5, color="#475569", linestyle=":", linewidth=2.0, label="Routine / non-routine split")
    ax.set_title(title, fontsize=11, fontweight="bold")
    ax.set_xlabel("Task ID", fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_xticks(x)
    ax.set_xticklabels(cost_df["task_id"].tolist())
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(fontsize=8, loc="best")
    fig.tight_layout()
    return fig


def build_automation_frontier_figure(cost_df: pd.DataFrame, metric: str = "unit_cost"):
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.8), sharey=True)
    mapping = [("routine", "Routine tasks"), ("non_routine", "Non-routine tasks")]
    if metric == "productivity_per_euro":
        low_column = "productivity_per_euro_low"
        high_column = "productivity_per_euro_high"
        ai_column = "productivity_per_euro_ai"
        ylabel = "Output per € (higher = better value)"
        ai_best_rule = lambda subset: subset["productivity_per_euro_ai"] >= subset[["productivity_per_euro_low", "productivity_per_euro_high"]].max(axis=1)
    else:
        low_column = "unit_cost_low"
        high_column = "unit_cost_high"
        ai_column = "unit_cost_ai"
        ylabel = "€ per unit output (lower = cheaper)"
        ai_best_rule = lambda subset: subset["unit_cost_ai"] <= subset[["unit_cost_low", "unit_cost_high"]].min(axis=1)
    for ax, (task_type, title) in zip(axes, mapping):
        subset = cost_df[cost_df["task_type"] == task_type].copy()
        if subset.empty:
            ax.set_title(title)
            ax.text(0.5, 0.5, "No tasks in this group", ha="center", va="center", transform=ax.transAxes)
            continue
        x = subset["complexity_index"]
        ax.plot(x, subset[low_column], label="Low-skill", color="#f97316", linewidth=2.0)
        ax.plot(x, subset[high_column], label="High-skill", color="#2563eb", linewidth=2.0)
        ax.plot(x, subset[ai_column], label="AI", color="#16a34a", linewidth=2.0)
        ai_best = ai_best_rule(subset)
        ax.scatter(x[ai_best], subset.loc[ai_best, ai_column], color="#16a34a", s=30, zorder=4)
        ax.set_title(title, fontsize=10, fontweight="bold")
        ax.set_xlabel("Task complexity index", fontsize=9)
        ax.grid(True, linestyle="--", alpha=0.5)
    axes[0].set_ylabel(ylabel, fontsize=9)
    axes[0].legend(fontsize=8, loc="best")
    fig.suptitle("Automation frontier plot", fontsize=11, fontweight="bold")
    fig.tight_layout()
    return fig


def _project_expected_wages(mode: str, params: Dict[str, float], assignment: Dict[str, float], t: int, adaptive_trend_high: float, adaptive_trend_low: float, mean_field_displacement_flow: float):
    mode = normalize_mode(mode)
    if mode == "npv_naive":
        return assignment["wage_high"], assignment["wage_low"]
    if mode == "npv_adaptive":
        return (
            max(params["a_h"], assignment["wage_high"] + t * adaptive_trend_high),
            max(params["a_l"], assignment["wage_low"] + t * adaptive_trend_low),
        )
    employed_high = assignment["employed_high"]
    employed_low = assignment["employed_low"]
    total = max(employed_high + employed_low, 1)
    share_high = employed_high / total
    share_low = employed_low / total
    proj_high = max(0.0, employed_high - t * mean_field_displacement_flow * share_high)
    proj_low = max(0.0, employed_low - t * mean_field_displacement_flow * share_low)
    return (
        params["a_h"] + params["b_h"] * proj_high,
        params["a_l"] + params["b_l"] * proj_low,
    )


def compute_example_npv_preview(
    params: Dict[str, float],
    alpha: float,
    task_type: str,
    complexity_index: int,
    mode: str,
    adaptive_trend_high: float = 0.0,
    adaptive_trend_low: float = 0.0,
    mean_field_displacement_flow: float = 0.0,
    wage_high_override: Optional[float] = None,
    wage_low_override: Optional[float] = None,
    ai_cost_override: Optional[float] = None,
    severance_tenure_years: float = 0.0,
    severance_n_workers: int = 1,
    severance_skill: str = "high",
    severance_rate_override: Optional[float] = None,
):
    mode = normalize_mode(mode)
    assignment = estimate_initial_assignment(alpha, params, n_tasks_per_producer=20)
    # Apply slider overrides so the waterfall / heatmap reflects the preview wages
    if wage_high_override is not None:
        assignment["wage_high"] = wage_high_override
    if wage_low_override is not None:
        assignment["wage_low"] = wage_low_override
    effective_k_ai = ai_cost_override if ai_cost_override is not None else params["k_ai"]
    complexity_index = max(1, complexity_index)
    productivity_low = _preview_productivity_human(
        params, task_type, complexity_index, "low"
    )
    productivity_high = _preview_productivity_human(
        params, task_type, complexity_index, "high"
    )
    productivity_ai = _preview_productivity_ai(params, task_type, complexity_index)

    investment_cost = params["I_base"] * (1.0 + params["complexity_scaling"] * complexity_index)
    discounted_human_cost = 0.0
    discounted_ai_cost = 0.0
    annual_rows = []

    for t in range(1, int(params["T_horizon"]) + 1):
        exp_w_h, exp_w_l = _project_expected_wages(
            mode,
            params,
            assignment,
            t,
            adaptive_trend_high,
            adaptive_trend_low,
            mean_field_displacement_flow,
        )
        uc_h = exp_w_h / max(productivity_high, 1e-9)
        uc_l = exp_w_l / max(productivity_low, 1e-9)
        human_unit_cost = min(uc_h, uc_l)
        future_k_ai = params["k_ai_floor"] + (effective_k_ai - params["k_ai_floor"]) * np.exp(-params["k_ai_decay"] * t)
        ai_unit_cost = future_k_ai / max(productivity_ai, 1e-9)
        discount = (1 + params["r_discount"]) ** t
        discounted_human_cost += human_unit_cost / discount
        discounted_ai_cost += ai_unit_cost / discount
        annual_rows.append(
            {
                "period": t,
                "expected_human_cost": human_unit_cost,
                "expected_ai_cost": ai_unit_cost,
                "discount_factor": discount,
                "discounted_saving": (human_unit_cost - ai_unit_cost) / discount,
            }
        )

    # Severance (transitievergoeding) component: only meaningful if workers are vast
    # and have tenure. We expose it as an upfront cost added to investment, mirroring
    # labour_market_model.py:496: NPV = -(I + severance) + discounted_savings.
    if severance_rate_override is not None:
        effective_sev_rate = float(severance_rate_override)
    else:
        effective_sev_rate = float(params.get("severance_rate", 0.0))
    if severance_skill == "low":
        wage_for_severance = float(assignment["wage_low"])
    else:
        wage_for_severance = float(assignment["wage_high"])
    n_workers = max(0, int(severance_n_workers))
    tenure_y = max(0.0, float(severance_tenure_years))
    severance_cost = effective_sev_rate * wage_for_severance * tenure_y * n_workers

    npv = -(investment_cost + severance_cost) + discounted_human_cost - discounted_ai_cost
    return {
        "npv": float(npv),
        "investment_cost": float(investment_cost),
        "severance_cost": float(severance_cost),
        "severance_rate_used": float(effective_sev_rate),
        "severance_tenure_years": float(tenure_y),
        "severance_n_workers": int(n_workers),
        "severance_skill": severance_skill,
        "severance_wage": float(wage_for_severance),
        "discounted_human_cost": float(discounted_human_cost),
        "discounted_ai_cost": float(discounted_ai_cost),
        "annual_table": pd.DataFrame(annual_rows),
        "productivity_low": float(productivity_low),
        "productivity_high": float(productivity_high),
        "productivity_ai": float(productivity_ai),
    }


def build_npv_waterfall_figure(npv_preview: Dict[str, float], mode_label: str, task_label: str):
    severance_cost = float(npv_preview.get("severance_cost", 0.0))
    labels = [
        "Upfront investment",
        "Severance (EP)",
        "PV human cost avoided",
        "PV AI cost incurred",
        "Net NPV",
    ]
    values = [
        -npv_preview["investment_cost"],
        -severance_cost,
        npv_preview["discounted_human_cost"],
        -npv_preview["discounted_ai_cost"],
        npv_preview["npv"],
    ]
    colors = ["#dc2626", "#ea580c", "#2563eb", "#16a34a", "#111827"]
    x = np.arange(len(labels))
    fig, ax = plt.subplots(figsize=(11, 4.8))
    ax.bar(x, values, color=colors)
    for idx, value in enumerate(values):
        ax.text(idx, value, f"{value:.2f}", ha="center", va="bottom" if value >= 0 else "top", fontsize=8)
    ax.axhline(0, color="#475569", linewidth=1.2)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=0)
    ax.set_ylabel("€ (present value)", fontsize=9)
    sev_subtitle = ""
    if severance_cost > 0:
        sev_subtitle = (
            f"  |  severance = {npv_preview.get('severance_rate_used', 0.0):.3f} "
            f"× €{npv_preview.get('severance_wage', 0.0):.2f} "
            f"× {npv_preview.get('severance_tenure_years', 0.0):.1f}y "
            f"× {npv_preview.get('severance_n_workers', 0)} workers "
            f"= €{severance_cost:.2f}"
        )
    ax.set_title(f"NPV waterfall - {mode_label} - {task_label}{sev_subtitle}", fontsize=11, fontweight="bold")
    ax.grid(True, axis="y", linestyle="--", alpha=0.5)
    fig.tight_layout()
    return fig


def build_npv_heatmap_figure(params: Dict[str, float], alpha: float, mode: str, adaptive_trend_high: float = 0.0, adaptive_trend_low: float = 0.0, mean_field_displacement_flow: float = 0.0, max_complexity: int = 20, wage_high_override: Optional[float] = None, wage_low_override: Optional[float] = None, ai_cost_override: Optional[float] = None):
    heatmap = np.full((2, max_complexity), np.nan)
    for row_idx, task_type in enumerate(["routine", "non_routine"]):
        for complexity in range(1, max_complexity + 1):
            preview = compute_example_npv_preview(
                params=params,
                alpha=alpha,
                task_type=task_type,
                complexity_index=complexity,
                mode=mode,
                adaptive_trend_high=adaptive_trend_high,
                adaptive_trend_low=adaptive_trend_low,
                mean_field_displacement_flow=mean_field_displacement_flow,
                wage_high_override=wage_high_override,
                wage_low_override=wage_low_override,
                ai_cost_override=ai_cost_override,
            )
            heatmap[row_idx, complexity - 1] = preview["npv"]
    fig, ax = plt.subplots(figsize=(12, 3.8))
    npv_cmap = LinearSegmentedColormap.from_list("npv_cmap", ["red", "white", "green"])
    norm = mcolors.TwoSlopeNorm(vmin=np.nanmin(heatmap), vcenter=0.0, vmax=np.nanmax(heatmap))
    im = ax.imshow(heatmap, aspect="auto", cmap=npv_cmap, norm=norm)
    ax.set_yticks([0, 1])
    ax.set_yticklabels(["Routine", "Non-routine"])
    ax.set_xticks(np.arange(max_complexity))
    ax.set_xticklabels([str(i) for i in range(1, max_complexity + 1)])
    ax.set_xlabel("Task complexity index", fontsize=9)
    ax.set_title(f"NPV heatmap - {MODE_LABELS.get(mode, mode)}", fontsize=11, fontweight="bold")
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("NPV (€)", fontsize=9)
    fig.tight_layout()
    return fig


def build_automation_complexity_figure(bundle: "RunBundle") -> plt.Figure:
    """
    Two-row automation complexity overview built from simulation results.

    Top row  – Timing scatter (one panel per NPV mode that has timing data):
        x = investment step, y = complexity index, colour = task type.
        Shows which tasks automated when and whether simple tasks lead.

    Bottom row – End-state heatmap (one panel per mode, including ULC):
        2 rows (Routine / Non-routine) × n_complexity columns.
        Cell colour = share of tasks at that slot that ended up automated.
        Cells are annotated with the percentage for quick reading.
    """
    timing_df = build_timing_dataframe(bundle)
    rng = np.random.default_rng(0)

    if not bundle.models:
        fig, ax = plt.subplots(figsize=(8, 2))
        ax.axis("off")
        ax.text(0.5, 0.5, "Automation complexity plot not available for loaded batches\n(model objects are not persisted — re-run to see this plot)",
                ha="center", va="center", fontsize=11, color="#6b7280", transform=ax.transAxes)
        fig.suptitle("Automation by task type and complexity", fontsize=13, fontweight="bold")
        return fig

    # --- Task universe from the first model (same structure across all modes) ---
    first_model = next(iter(bundle.models.values()))
    task_rows = [
        {"task_type": t.task_type, "complexity_index": t.complexity_index}
        for p in first_model.producers
        for t in p.tasks
    ]
    all_tasks_df = pd.DataFrame(task_rows)
    routine_df = all_tasks_df[all_tasks_df["task_type"] == "routine"]
    nr_df = all_tasks_df[all_tasks_df["task_type"] == "non_routine"]
    max_r_c = int(routine_df["complexity_index"].max()) if not routine_df.empty else 10
    max_nr_c = int(nr_df["complexity_index"].max()) if not nr_df.empty else 10
    n_complexity = max(max_r_c, max_nr_c)

    npv_modes_present = [m for m in bundle.modes if m in NPV_MODES and not timing_df.empty and m in timing_df["mode"].values]
    all_modes = bundle.modes
    n_npv = len(npv_modes_present)
    n_all = len(all_modes)
    has_scatter = n_npv > 0

    # --- Build end-state heatmap arrays ---
    heatmaps: Dict[str, np.ndarray] = {}
    for mode in all_modes:
        if mode not in bundle.models:
            continue
        model = bundle.models[mode]
        data_r: Dict[int, list] = {}
        data_nr: Dict[int, list] = {}
        for p in model.producers:
            for t in p.tasks:
                bucket = data_r if t.task_type == "routine" else data_nr
                bucket.setdefault(t.complexity_index, []).append(int(t.automated))
        hmap = np.full((2, n_complexity), np.nan)
        for c in range(1, max_r_c + 1):
            vals = data_r.get(c, [])
            hmap[0, c - 1] = float(np.mean(vals)) if vals else 0.0
        for c in range(1, max_nr_c + 1):
            vals = data_nr.get(c, [])
            hmap[1, c - 1] = float(np.mean(vals)) if vals else 0.0
        heatmaps[mode] = hmap

    # --- Figure layout ---
    n_rows = (1 if has_scatter else 0) + 1
    n_cols = max(n_npv if has_scatter else 0, n_all)
    fig = plt.figure(figsize=(max(10, 4 * n_cols), 5 * n_rows))
    gs = gridspec.GridSpec(
        n_rows, n_cols,
        figure=fig,
        hspace=0.55,
        wspace=0.38,
    )
    fig.suptitle(
        "Automation by task type and complexity",
        fontsize=13,
        fontweight="bold",
        y=1.01,
    )

    # --- Row 0: timing scatter ---
    if has_scatter:
        type_colors = {"routine": "#f97316", "non_routine": "#2563eb"}
        type_labels = {"routine": "Routine", "non_routine": "Non-routine"}
        for col_idx, mode in enumerate(npv_modes_present):
            ax = fig.add_subplot(gs[0, col_idx])
            mode_timing = timing_df[timing_df["mode"] == mode]
            for task_type, color in type_colors.items():
                sub = mode_timing[mode_timing["task_type"] == task_type]
                if sub.empty:
                    continue
                jitter = rng.uniform(-0.28, 0.28, size=len(sub))
                ax.scatter(
                    sub["investment_step"],
                    sub["complexity_index"] + jitter,
                    c=color,
                    s=14,
                    alpha=0.55,
                    linewidths=0,
                    label=type_labels[task_type],
                    zorder=3,
                )
            ax.set_xlabel("Automation step", fontsize=8)
            ax.set_ylabel("Complexity index", fontsize=8)
            ax.set_title(
                f"Timing — {MODE_LABELS.get(mode, mode)}",
                fontsize=9,
                fontweight="bold",
                pad=6,
            )
            ax.set_xlim(-2, bundle.n_steps + 2)
            ax.set_ylim(0.3, n_complexity + 0.7)
            ax.set_yticks(range(1, n_complexity + 1))
            ax.grid(True, linestyle="--", alpha=0.35)
            ax.tick_params(labelsize=7)
            ax.legend(fontsize=6.5, loc="upper left", markerscale=1.4)

    # --- Row 1 (or 0): automation frontier line plots (2 panels: routine | non-routine) ---
    # Each panel shows one line per mode: x = complexity index, y = share automated.
    # This makes the automation gradient and cutoff point immediately readable.
    frontier_row = 1 if has_scatter else 0
    n_frontier_cols = max(n_cols, 2)
    # Re-build gridspec with 2 columns for the frontier row if needed
    # We use subplot2grid-style spans: split the frontier row into 2 equal panels
    # regardless of n_cols (scatter row may have 1–3 panels)
    complexity_indices = list(range(1, n_complexity + 1))

    type_panel_spec = [
        ("routine",     "Routine tasks",     max_r_c),
        ("non_routine", "Non-routine tasks", max_nr_c),
    ]

    # Place the two frontier panels spanning equal halves of the bottom row
    half = n_cols // 2 if n_cols >= 2 else 1
    panel_col_spans = [
        (0, max(1, half)),
        (max(1, half), n_cols),
    ]
    # If n_cols == 1 (only one mode, no scatter), fall back to separate subplots
    if n_cols < 2:
        gs_frontier = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[frontier_row, 0], wspace=0.4)
        frontier_axes = [fig.add_subplot(gs_frontier[0, 0]), fig.add_subplot(gs_frontier[0, 1])]
    else:
        frontier_axes = [
            fig.add_subplot(gs[frontier_row, panel_col_spans[0][0]:panel_col_spans[0][1]]),
            fig.add_subplot(gs[frontier_row, panel_col_spans[1][0]:panel_col_spans[1][1]]),
        ]

    for ax, (task_type, panel_title, max_c) in zip(frontier_axes, type_panel_spec):
        x = list(range(1, max_c + 1))
        for mode in all_modes:
            hmap = heatmaps[mode]
            row_idx = 0 if task_type == "routine" else 1
            y = [hmap[row_idx, c - 1] for c in x]
            ax.plot(
                x, y,
                label=MODE_LABELS.get(mode, mode),
                color=MODE_COLORS.get(mode, "#555"),
                linestyle=MODE_LS.get(mode, "-"),
                marker=MODE_MARKERS.get(mode, "o"),
                markersize=5,
                linewidth=2.0,
            )
        ax.set_xlim(0.5, max_c + 0.5)
        ax.set_ylim(-0.04, 1.08)
        ax.set_xticks(range(1, max_c + 1))
        ax.set_xticklabels([str(i) for i in range(1, max_c + 1)], fontsize=7)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
        ax.set_xlabel("Complexity index  (1 = simple → higher = more complex)", fontsize=8)
        ax.set_ylabel("Share of tasks automated", fontsize=8)
        ax.set_title(panel_title, fontsize=10, fontweight="bold", pad=8)
        ax.tick_params(labelsize=7)
        ax.grid(True, linestyle="--", alpha=0.4)
        ax.axhline(0, color="#94a3b8", linewidth=0.6, linestyle=":")
        ax.axhline(1, color="#94a3b8", linewidth=0.6, linestyle=":")
        ax.legend(fontsize=7.5, loc="upper right")

    fig.tight_layout()
    return fig


# =============================================================================
# New figure: alpha bins × task status (categorical, last timestep)
# =============================================================================

def build_alpha_status_figure(bundle: "RunBundle") -> plt.Figure:
    """
    For each adoption mode, classify every task at the last timestep into one of
    five categories (AI / High-skill / Low-skill / Mixed / Empty) and show the
    count per fixed alpha bin (width 0.05 from 0→1, 20 bins total).

    One subplot per adoption mode.  Tasks are counted, not worker-slots.
    """
    if not bundle.models:
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.axis("off")
        ax.text(0.5, 0.5,
                "Model objects not available — re-run the simulation to generate this plot.",
                ha="center", va="center", fontsize=11, color="#6b7280",
                transform=ax.transAxes)
        fig.suptitle("Task status by alpha range (last timestep)", fontsize=13, fontweight="bold")
        return fig

    modes = [m for m in bundle.modes if m in bundle.models]
    n_modes = len(modes)
    fig, axes = plt.subplots(1, n_modes, figsize=(max(8, 6 * n_modes), 5), squeeze=False)

    bin_edges = np.arange(0.0, 1.0 + 1e-9, 0.05)  # 21 edges → 20 bins
    status_colors = {
        "AI":         "#4c72b0",
        "High-skill": "#55a868",
        "Low-skill":  "#dd8452",
        "Mixed":      "#9467bd",
        "Empty":      "#d3d3d3",
    }
    statuses = ["AI", "High-skill", "Low-skill", "Mixed", "Empty"]

    for ax, mode in zip(axes[0], modes):
        model = bundle.models[mode]
        records = []
        for p in model.producers:
            task_counts = {s: 0 for s in statuses}
            for t in p.tasks:
                if t.automated:
                    status = "AI"
                else:
                    has_high = any(w.skill_level == "high" for w in t.employees)
                    has_low  = any(w.skill_level == "low"  for w in t.employees)
                    if has_high and has_low:
                        status = "Mixed"
                    elif has_high:
                        status = "High-skill"
                    elif has_low:
                        status = "Low-skill"
                    else:
                        status = "Empty"
                task_counts[status] += 1
            rec = {"alpha": p.alpha}
            rec.update(task_counts)
            records.append(rec)

        df = pd.DataFrame(records)
        df["bin"] = pd.cut(df["alpha"], bins=bin_edges, include_lowest=True)
        binned = df.groupby("bin", observed=False)[statuses].sum()

        x = np.arange(len(binned))
        bottoms = np.zeros(len(binned))
        for s in statuses:
            vals = binned[s].to_numpy(dtype=float)
            ax.bar(x, vals, bottom=bottoms, label=s,
                   color=status_colors[s], alpha=0.88, width=0.7)
            bottoms += vals

        tick_labels = [
            f"{interval.left:.2f}" if i % 2 == 0 else ""
            for i, interval in enumerate(binned.index)
        ]
        ax.set_xticks(x)
        ax.set_xticklabels(tick_labels, rotation=45, ha="right", fontsize=7.5)
        ax.set_xlabel("Alpha  (routine-task share, bins of 0.05)", fontsize=9)
        ax.set_ylabel("Number of tasks", fontsize=9)
        ax.set_title(MODE_LABELS.get(mode, mode), fontsize=10, fontweight="bold", pad=8)
        ax.tick_params(labelsize=8)
        ax.grid(True, axis="y", linestyle="--", alpha=0.4)
        ax.legend(fontsize=8, loc="upper right")

    fig.suptitle(
        "Task status by alpha range — last simulation step",
        fontsize=12, fontweight="bold", y=1.01,
    )
    fig.tight_layout()
    return fig


# =============================================================================
# New figure: alpha bins × workforce composition (stacked bar, last timestep)
# =============================================================================

def build_alpha_workforce_figure(bundle: "RunBundle") -> plt.Figure:
    """
    For each adoption mode, show how the workforce inside firms is distributed
    across alpha (routine-task share) ranges at the final simulation timestep.

    X-axis : fixed bins of width 0.05 spanning 0.00 → 1.00 (20 bins total).
             Bins with no firms are shown as empty bars.
    Y-axis : total worker-slots occupied by AI, high-skill, or low-skill workers
             summed over all firms in that alpha bin.
    One subplot per adoption mode.
    """
    if not bundle.models:
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.axis("off")
        ax.text(0.5, 0.5,
                "Model objects not available — re-run the simulation to generate this plot.",
                ha="center", va="center", fontsize=11, color="#6b7280",
                transform=ax.transAxes)
        fig.suptitle("Workforce by alpha range (last timestep)", fontsize=13, fontweight="bold")
        return fig

    modes = [m for m in bundle.modes if m in bundle.models]
    n_modes = len(modes)
    fig, axes = plt.subplots(1, n_modes, figsize=(max(8, 6 * n_modes), 5), squeeze=False)

    # Fixed bins: 0.00, 0.05, 0.10, …, 1.00  → 20 bins of width 0.05
    bin_edges = np.arange(0.0, 1.0 + 1e-9, 0.05)  # 21 edges → 20 bins

    colors = {"AI": "#4c72b0", "High-skill": "#55a868", "Low-skill": "#dd8452"}
    keys = ["AI", "High-skill", "Low-skill"]

    for ax, mode in zip(axes[0], modes):
        model = bundle.models[mode]
        records = []
        for p in model.producers:
            n_ai   = sum(t.n_ai for t in p.tasks if t.automated)
            n_high = sum(1 for t in p.tasks for w in t.employees if w.skill_level == "high")
            n_low  = sum(1 for t in p.tasks for w in t.employees if w.skill_level == "low")
            records.append({"alpha": p.alpha, "AI": n_ai, "High-skill": n_high, "Low-skill": n_low})
        df = pd.DataFrame(records)
        df["bin"] = pd.cut(df["alpha"], bins=bin_edges, include_lowest=True)
        binned = df.groupby("bin", observed=False)[keys].sum()

        x = np.arange(len(binned))
        bottoms = np.zeros(len(binned))
        for k in keys:
            vals = binned[k].to_numpy(dtype=float)
            ax.bar(x, vals, bottom=bottoms, label=k, color=colors[k], alpha=0.88, width=0.7)
            bottoms += vals

        # Label every other tick to avoid crowding (0.00, 0.10, 0.20, …)
        tick_labels = [
            f"{interval.left:.2f}" if i % 2 == 0 else ""
            for i, interval in enumerate(binned.index)
        ]
        ax.set_xticks(x)
        ax.set_xticklabels(tick_labels, rotation=45, ha="right", fontsize=7.5)
        ax.set_xlabel("Alpha  (routine-task share, bins of 0.05)", fontsize=9)
        ax.set_ylabel("Total worker-slots", fontsize=9)
        ax.set_title(MODE_LABELS.get(mode, mode), fontsize=10, fontweight="bold", pad=8)
        ax.tick_params(labelsize=8)
        ax.grid(True, axis="y", linestyle="--", alpha=0.4)
        ax.legend(fontsize=8, loc="upper right")

    fig.suptitle(
        "Workforce composition by alpha range — last simulation step",
        fontsize=12, fontweight="bold", y=1.01,
    )
    fig.tight_layout()
    return fig


# =============================================================================
# Figure: Task status by complexity (categorical: AI / High / Low / Mixed / Empty)
# =============================================================================

def build_task_complexity_status_figure(bundle: "RunBundle") -> plt.Figure:
    """
    For each adoption mode (rows) and task type (routine / non-routine),
    classify every task at the last timestep into one of five mutually-exclusive
    categories and display the count per complexity level (1–max_c).

    Categories
    ----------
    AI        : task is automated (t.automated == True)
    High-skill: at least one high-skill worker, no low-skill workers
    Low-skill : at least one low-skill worker, no high-skill workers
    Mixed     : both high-skill and low-skill workers present
    Empty     : no workers and not automated
    """
    if not bundle.models:
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.axis("off")
        ax.text(0.5, 0.5,
                "Model objects not available — re-run the simulation to generate this plot.",
                ha="center", va="center", fontsize=11, color="#6b7280",
                transform=ax.transAxes)
        fig.suptitle("Task status by complexity (last timestep)", fontsize=13, fontweight="bold")
        return fig

    modes = [m for m in bundle.modes if m in bundle.models]
    n_modes = len(modes)

    max_c = max(
        t.complexity_index
        for m in modes
        for p in bundle.models[m].producers
        for t in p.tasks
    )
    x = np.arange(1, max_c + 1)

    status_colors = {
        "AI":         "#4c72b0",
        "High-skill": "#55a868",
        "Low-skill":  "#dd8452",
        "Mixed":      "#9467bd",
        "Empty":      "#d3d3d3",
    }
    statuses = ["AI", "High-skill", "Low-skill", "Mixed", "Empty"]
    type_labels = {"routine": "Routine tasks", "non_routine": "Non-routine tasks"}

    fig, axes = plt.subplots(
        n_modes, 2,
        figsize=(13, max(4, 3.5 * n_modes)),
        squeeze=False,
    )

    for row_idx, mode in enumerate(modes):
        model = bundle.models[mode]

        # Count tasks per (task_type, complexity_index, status)
        counts: dict = {}
        for p in model.producers:
            for t in p.tasks:
                if t.automated:
                    status = "AI"
                else:
                    has_high = any(w.skill_level == "high" for w in t.employees)
                    has_low  = any(w.skill_level == "low"  for w in t.employees)
                    if has_high and has_low:
                        status = "Mixed"
                    elif has_high:
                        status = "High-skill"
                    elif has_low:
                        status = "Low-skill"
                    else:
                        status = "Empty"
                key = (t.task_type, t.complexity_index, status)
                counts[key] = counts.get(key, 0) + 1

        for col_idx, task_type in enumerate(["routine", "non_routine"]):
            ax = axes[row_idx][col_idx]
            bottoms = np.zeros(max_c)
            for s in statuses:
                vals = np.array([
                    counts.get((task_type, c, s), 0)
                    for c in range(1, max_c + 1)
                ], dtype=float)
                ax.bar(x, vals, bottom=bottoms, label=s,
                       color=status_colors[s], alpha=0.88, width=0.8)
                bottoms += vals

            ax.set_xticks(x)
            ax.set_xticklabels([str(c) for c in x], fontsize=7)
            ax.set_xlabel("Complexity index  (1 = simplest)", fontsize=8.5)
            ax.set_ylabel("Number of tasks", fontsize=8.5)
            panel_title = f"{MODE_LABELS.get(mode, mode)}  —  {type_labels[task_type]}"
            ax.set_title(panel_title, fontsize=9.5, fontweight="bold", pad=7)
            ax.tick_params(labelsize=8)
            ax.grid(True, axis="y", linestyle="--", alpha=0.4)
            ax.legend(fontsize=8, loc="upper right")

    fig.suptitle(
        "Task status by complexity — last simulation step",
        fontsize=12, fontweight="bold", y=1.01,
    )
    fig.tight_layout()
    return fig


# =============================================================================
# New figure: task complexity × task type × workforce (stacked bar, last timestep)
# =============================================================================

def build_task_complexity_workforce_figure(bundle: "RunBundle") -> plt.Figure:
    """
    For each adoption mode (rows), show two panels (routine / non-routine).
    X-axis: complexity index 1..20.
    Y-axis: absolute number of worker-slots (AI units + high-skill + low-skill workers)
            summed across all tasks of that complexity in that mode.
    Bar height reflects both how many tasks exist at that complexity level and
    how intensively each task is staffed.
    """
    if not bundle.models:
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.axis("off")
        ax.text(0.5, 0.5,
                "Model objects not available — re-run the simulation to generate this plot.",
                ha="center", va="center", fontsize=11, color="#6b7280",
                transform=ax.transAxes)
        fig.suptitle("Absolute worker count by task complexity (last timestep)", fontsize=13, fontweight="bold")
        return fig

    modes = [m for m in bundle.modes if m in bundle.models]
    n_modes = len(modes)

    fig, axes = plt.subplots(
        n_modes, 2,
        figsize=(13, max(4, 3.5 * n_modes)),
        squeeze=False,
    )

    colors = {"AI": "#4c72b0", "High-skill": "#55a868", "Low-skill": "#dd8452"}
    keys = ["AI", "High-skill", "Low-skill"]
    type_labels = {"routine": "Routine tasks", "non_routine": "Non-routine tasks"}

    max_c = max(
        t.complexity_index
        for m in modes
        for p in bundle.models[m].producers
        for t in p.tasks
    )
    x = np.arange(1, max_c + 1)

    for row_idx, mode in enumerate(modes):
        model = bundle.models[mode]

        # Aggregate worker-slots per (task_type, complexity_index)
        data: dict = {}
        for p in model.producers:
            for t in p.tasks:
                dk = (t.task_type, t.complexity_index)
                if dk not in data:
                    data[dk] = {"AI": 0, "High-skill": 0, "Low-skill": 0}
                if t.automated:
                    data[dk]["AI"] += t.n_ai
                else:
                    for w in t.employees:
                        sk = "High-skill" if w.skill_level == "high" else "Low-skill"
                        data[dk][sk] += 1

        for col_idx, task_type in enumerate(["routine", "non_routine"]):
            ax = axes[row_idx][col_idx]
            bottoms = np.zeros(max_c)
            for k in keys:
                vals = np.array([
                    data.get((task_type, c), {}).get(k, 0)
                    for c in range(1, max_c + 1)
                ], dtype=float)
                ax.bar(x, vals, bottom=bottoms, label=k, color=colors[k], alpha=0.88, width=0.8)
                bottoms += vals

            ax.set_xticks(x)
            ax.set_xticklabels([str(c) for c in x], fontsize=7)
            ax.set_xlabel("Complexity index  (1 = simplest)", fontsize=8.5)
            ax.set_ylabel("Total worker-slots", fontsize=8.5)
            panel_title = f"{MODE_LABELS.get(mode, mode)}  —  {type_labels[task_type]}"
            ax.set_title(panel_title, fontsize=9.5, fontweight="bold", pad=7)
            ax.tick_params(labelsize=8)
            ax.grid(True, axis="y", linestyle="--", alpha=0.4)
            ax.legend(fontsize=8, loc="upper right")

    fig.suptitle(
        "Absolute worker count by task complexity — last simulation step",
        fontsize=12, fontweight="bold", y=1.01,
    )
    fig.tight_layout()
    return fig
